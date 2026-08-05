"""Regression tests for the Summary and Weekly sheet builders.

Unlike test_jira_report.py, these tests need a *real* openpyxl (the sheet
builders call actual Workbook/Worksheet APIs — DataValidation, conditional
formatting, cell styling — that a stub can't stand in for). Run with the
project virtualenv, which has real openpyxl installed:

    .venv/bin/python -m pytest tests/test_report_sheets.py -q
"""
import sys
import types
import importlib.util
import datetime as dt
from pathlib import Path

import pytest

# Discard any openpyxl stub test_jira_report.py may have installed in
# sys.modules under the same pytest run, then do ONE fresh real import that
# both this file and the freshly-loaded jira-report.py module will share.
# (Importing openpyxl separately before *and* after clearing sys.modules
# would bind two different Side/Border class generations — objects that
# print identically but fail isinstance() checks against each other.)
_saved_stubs = {name: sys.modules.pop(name) for name in list(sys.modules) if name == "openpyxl" or name.startswith("openpyxl.")}
try:
    import openpyxl
except ImportError:
    sys.modules.update(_saved_stubs)
    pytest.skip("real openpyxl required — run via .venv", allow_module_level=True)

if "google_drive_sync" not in sys.modules:
    _gds = types.ModuleType("google_drive_sync")
    _gds.upload_or_update = lambda *a, **kw: {}
    sys.modules["google_drive_sync"] = _gds

APP_DIR = Path(__file__).resolve().parents[1] / "app"
spec = importlib.util.spec_from_file_location("jr_sheets", APP_DIR / "jira-report.py")
jr = importlib.util.module_from_spec(spec)
spec.loader.exec_module(jr)

# Deliberately NOT restoring _saved_stubs into sys.modules here: openpyxl's
# own internals do lazy `from openpyxl.x import y` imports inside methods
# like Workbook.save() — those resolve against whatever sys.modules holds
# at CALL time, not at jira-report.py's import time. Putting the stub back
# would make any later test in this file that actually writes a file (via
# build_xlsx) crash inside openpyxl's own save() with an unrelated-looking
# import error. No other test file depends on sys.modules["openpyxl"]
# being a stub — they all mock at the function level (e.g.
# patch.object(rdu, "_openpyxl_importable", ...)) instead.

REPORT_DATE = dt.date(2026, 7, 21)


def _row(feature, link, status, start=None, end=None, task_type="Task",
         substream="", created_date=None, eta="", delta_eta="", seq=0, epic_summary=""):
    """Build one lifecycle row the way issue_rows() would."""
    return {
        "Feature": feature,
        "Epic": f"{link}-EPIC",
        "Epic Summary": epic_summary,
        "Substream": substream,
        "Task type": task_type,
        "Task": f"task {link}",
        "Status": status,
        "Start": start,
        "End": end,
        "Done week": (end.isocalendar()[1] if status in ("done", "rejected") and end else ""),
        "ETA": eta,
        "Days in Work": "",
        "Delta ETA": delta_eta,
        "Link": link,
        "Created date": created_date or start,
        "Created week": ((created_date or start).isocalendar()[1] if (created_date or start) else ""),
        "_seq": seq,
    }


def _new_sheets():
    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "summary"
    weekly = wb.create_sheet("weekly")
    return wb, summary, weekly


def _summary_row_dict(sheet, feature):
    # Layout: row1 label, row2 date, row3 group headers, row4 column
    # headers, row5+ data.
    headers = [sheet.cell(row=4, column=c).value for c in range(1, sheet.max_column + 1)]
    for r in range(5, sheet.max_row + 1):
        if sheet.cell(row=r, column=1).value == feature:
            return {h: sheet.cell(row=r, column=i + 1).value for i, h in enumerate(headers)}
    raise AssertionError(f"feature {feature!r} not found in summary sheet")


def _weekly_data_rows(wb, feature=None):
    wd = wb["_weekly_data"]
    rows = [[wd.cell(row=r, column=c).value for c in range(1, 11)] for r in range(2, wd.max_row + 1)]
    if feature is not None:
        rows = [r for r in rows if r[0] == feature]
    return rows


# ── Summary: Actual TTM unions parallel work, doesn't sum it ────────────────

def test_summary_actual_ttm_unions_parallel_tasks_not_sums():
    # Two tasks both worked the exact same 5 business days in parallel.
    start, end = dt.date(2026, 1, 5), dt.date(2026, 1, 9)  # Mon-Fri
    rows = [
        _row("F1", "A-1", "done", start=start, end=end),
        _row("F1", "A-2", "done", start=start, end=end),
    ]
    wb, summary, _ = _new_sheets()
    jr.build_summary_sheet(summary, rows, {}, {}, REPORT_DATE, {})
    result = _summary_row_dict(summary, "F1")
    # 5 business days ≈ 5 / 21.74 months ≈ 0.2 — NOT 10 (which summing would give).
    assert result["Actual"] == pytest.approx(0.2, abs=0.05)


# ── Summary: on-hold row date range counts as real active work ─────────────

def test_summary_on_hold_row_counts_as_active_work():
    # Task worked Mon-Fri, then flips to on hold with that exact range, then
    # later (after a gap) finishes as done with a second work burst.
    hold_start, hold_end = dt.date(2026, 1, 5), dt.date(2026, 1, 9)
    done_start, done_end = dt.date(2026, 1, 19), dt.date(2026, 1, 21)
    rows = [
        _row("F1", "A-1", "on hold", start=hold_start, end=hold_end, seq=0),
        _row("F1", "A-1", "done", start=done_start, end=done_end, seq=1),
    ]
    wb, summary, _ = _new_sheets()
    jr.build_summary_sheet(summary, rows, {}, {}, REPORT_DATE, {})
    result = _summary_row_dict(summary, "F1")
    # 5 (on-hold burst) + 3 (done burst) = 8 active business days ≈ 0.37mo.
    assert result["Actual"] == pytest.approx(8 / (30.4375 * 5 / 7), abs=0.05)
    # Hold time = the gap between the two bursts (Jan12-16, 5 business days).
    assert result["Hold, mo"] == pytest.approx(5 / (30.4375 * 5 / 7), abs=0.05)


# ── Summary: pace/ETA-from-actual excludes on-hold-only weeks ───────────────

def test_summary_pace_excludes_on_hold_only_weeks():
    # Feature with a done task and, in a totally different (much later) week,
    # a still-open on-hold task. The on-hold week must not dilute pace.
    done_start, done_end = dt.date(2026, 1, 5), dt.date(2026, 1, 9)
    hold_start, hold_end = dt.date(2026, 5, 4), dt.date(2026, 5, 8)
    rows = [
        _row("F1", "A-1", "done", start=done_start, end=done_end),
        _row("F1", "A-2", "on hold", start=hold_start, end=hold_end),
    ]
    wb, summary, _ = _new_sheets()
    jr.build_summary_sheet(summary, rows, {}, {}, REPORT_DATE, {})
    result = _summary_row_dict(summary, "F1")
    # Only the done task's single week should count toward pace, so
    # Actual pace = 1 done / 1 active(done) week = 1, not 1/2.
    assert result["Actual pace"] == 1


# ── Summary: Δ ETA, mo sign matches early/late delivery ─────────────────────

def test_summary_delta_eta_sign_late_and_early():
    eta_dates = {"Late": dt.date(2026, 1, 1), "Early": dt.date(2026, 3, 1)}
    rows = [
        _row("Late", "A-1", "done", start=dt.date(2026, 1, 5), end=dt.date(2026, 1, 20)),
        _row("Early", "B-1", "done", start=dt.date(2026, 1, 5), end=dt.date(2026, 1, 20)),
    ]
    wb, summary, _ = _new_sheets()
    jr.build_summary_sheet(summary, rows, {}, eta_dates, REPORT_DATE, {})
    late = _summary_row_dict(summary, "Late")
    early = _summary_row_dict(summary, "Early")
    assert late["Δ ETA, mo"] > 0     # delivered after its ETA
    assert early["Δ ETA, mo"] < 0    # delivered before its ETA


# ── Summary: no configured ETA/pace → ETA and Δ ETA stay blank, even done ──

def test_summary_no_eta_configured_stays_blank_even_when_done():
    # Fully done feature, but no ETA and no expected pace were ever set for it.
    rows = [_row("F1", "A-1", "done", start=dt.date(2026, 1, 5), end=dt.date(2026, 1, 9))]
    wb, summary, _ = _new_sheets()
    jr.build_summary_sheet(summary, rows, {}, {}, REPORT_DATE, {})
    result = _summary_row_dict(summary, "F1")
    assert result["ETA"] in (None, "")
    assert result["Δ ETA, mo"] in (None, "")
    assert result["Delivered"] is not None  # the real answer still shows up here


# ── Summary: blank spacer row between the date and the rest of the table ───

def test_summary_date_is_alone_on_its_own_row():
    rows = [_row("F1", "A-1", "done", start=dt.date(2026, 1, 5), end=dt.date(2026, 1, 9))]
    wb, summary, _ = _new_sheets()
    jr.build_summary_sheet(summary, rows, {}, {}, REPORT_DATE, {})
    assert summary["A1"].value == "Updated at"
    assert summary["A2"].value is not None            # the date, alone on row 2
    assert summary["B2"].value is None                # no group labels sharing the date's row
    assert summary["A3"].value is None                # Feature has no group label
    assert summary["B3"].value == "Pace - tasks per week (done+rejected)"  # group labels on row 3
    assert summary["A4"].value == "Feature"                                # column headers
    assert summary["A5"].value == "F1"                                     # data
    assert summary.freeze_panes == "B5"  # rows 1-4 AND column A (Feature) frozen


# ── Weekly: only weeks with an actual closure get a row ─────────────────────

def test_weekly_hides_created_only_weeks():
    rows = [
        # Closed in week of Jan 5, 2026.
        _row("F1", "A-1", "done", start=dt.date(2026, 1, 5), end=dt.date(2026, 1, 9),
             created_date=dt.date(2026, 1, 5)),
        # Created much later (week of May 4) but never closed — must not
        # produce its own weekly row.
        _row("F1", "A-2", "in progress", start=dt.date(2026, 5, 4), end=None,
             created_date=dt.date(2026, 5, 4)),
    ]
    wb, _, weekly = _new_sheets()
    jr.build_weekly_sheet(weekly, rows, REPORT_DATE, {}, feature_names=["F1"])
    weeks = _weekly_data_rows(wb, "F1")
    assert len(weeks) == 1
    assert weeks[0][1] == dt.date(2026, 1, 5).isocalendar()[1]


# ── Weekly: "All" sums correctly across features, no duplicate weeks ───────

def test_weekly_all_sums_across_features_no_duplicates():
    same_week_start = dt.date(2026, 1, 5)
    same_week_end = dt.date(2026, 1, 9)
    rows = [
        _row("F1", "A-1", "done", start=same_week_start, end=same_week_end),
        _row("F2", "B-1", "done", start=same_week_start, end=same_week_end),
        _row("F2", "B-2", "rejected", start=same_week_start, end=same_week_end),
    ]
    wb, _, weekly = _new_sheets()
    jr.build_weekly_sheet(weekly, rows, REPORT_DATE, {}, feature_names=["F1", "F2"])
    all_weeks = _weekly_data_rows(wb, "All")
    assert len(all_weeks) == 1  # one shared week, not two
    _, wn, _, tasks_done, tasks_rejected, bugs_done, bugs_rejected, total, created, _ = all_weeks[0]
    assert (tasks_done, tasks_rejected) == (2, 1)  # F1 done + F2 done, F2 rejected
    assert total == 3


# ── Weekly: "All" epic-name fallback doesn't leak the label "All" ──────────

def test_weekly_all_epic_name_does_not_leak():
    rows = [
        _row("F1", "A-1", "done", start=dt.date(2026, 1, 5), end=dt.date(2026, 1, 9), substream=""),
    ]
    wb, _, weekly = _new_sheets()
    jr.build_weekly_sheet(weekly, rows, REPORT_DATE, {}, feature_names=["F1"])
    ed = wb["_epic_data"]
    names = [ed.cell(row=r, column=3).value for r in range(2, ed.max_row + 1) if ed.cell(row=r, column=1).value == "All"]
    assert names == ["F1"]
    assert "All" not in names


# ── Weekly: epic name prefers the real Jira summary over Substream ─────────

def test_weekly_epic_name_prefers_real_summary():
    rows = [
        _row("F1", "A-1", "done", start=dt.date(2026, 1, 5), end=dt.date(2026, 1, 9),
             substream="Trimmed substream", epic_summary="Full original epic summary"),
    ]
    wb, _, weekly = _new_sheets()
    jr.build_weekly_sheet(weekly, rows, REPORT_DATE, {}, feature_names=["F1"])
    ed = wb["_epic_data"]
    names = [ed.cell(row=r, column=3).value for r in range(2, ed.max_row + 1) if ed.cell(row=r, column=1).value == "F1"]
    assert names == ["Full original epic summary"]


def test_weekly_epic_name_falls_back_without_summary():
    # No Epic Summary captured (e.g. older data) — falls back to Substream.
    rows = [
        _row("F1", "A-1", "done", start=dt.date(2026, 1, 5), end=dt.date(2026, 1, 9),
             substream="Trimmed substream"),
    ]
    wb, _, weekly = _new_sheets()
    jr.build_weekly_sheet(weekly, rows, REPORT_DATE, {}, feature_names=["F1"])
    ed = wb["_epic_data"]
    names = [ed.cell(row=r, column=3).value for r in range(2, ed.max_row + 1) if ed.cell(row=r, column=1).value == "F1"]
    assert names == ["Trimmed substream"]


# ── Weekly: Required moved next to the dropdown (row 2), rows 3-4 removed ──

def test_weekly_required_sits_on_row2_with_dropdown():
    rows = [_row("F1", "A-1", "done", start=dt.date(2026, 1, 5), end=dt.date(2026, 1, 9))]
    wb, _, weekly = _new_sheets()
    jr.build_weekly_sheet(weekly, rows, REPORT_DATE, {}, expected_tasks_per_week={},
                           feature_eta_dates={"F1": dt.date(2026, 1, 1)}, feature_names=["F1"])
    assert weekly["A2"].value == "All"         # dropdown defaults to the synthetic "All" entry
    assert weekly["F2"].value == "Required"    # Required label, same row as dropdown
    assert weekly["A3"].value is None          # spacer row, nothing left over
    assert weekly["A4"].value == "Weekly progress"  # group headers now on row 4
    assert weekly["A5"].value == "Week start"       # column headers now on row 5


# ── Weekly: sorts by real calendar date across year boundaries ─────────────

def test_weekly_sorts_by_actual_date_across_years():
    rows = [
        _row("F1", "A-1", "done", start=dt.date(2025, 12, 22), end=dt.date(2025, 12, 26)),  # ISO week 52, 2025
        _row("F1", "A-2", "done", start=dt.date(2026, 1, 5), end=dt.date(2026, 1, 9)),        # ISO week 2, 2026
    ]
    wb, _, weekly = _new_sheets()
    jr.build_weekly_sheet(weekly, rows, REPORT_DATE, {}, feature_names=["F1"])
    weeks = _weekly_data_rows(wb, "F1")
    weeks.sort(key=lambda r: int(r[9].split("|")[1]))  # seq order, as written
    dates = [r[2] for r in weeks]
    assert dates == sorted(dates, reverse=True)  # newest first
    assert dates[0] > dates[1]  # Jan 2026 week sorts above Dec 2025 week


# ── Weekly: header row is frozen so it stays visible on scroll ─────────────

def test_weekly_freezes_header_row():
    # Layout: row1 label, row2 dropdown+Required, row3 spacer, row4 group
    # headers, row5 column headers, row6+ data — freeze must cover through
    # row5 so DATA_START (row6) is the first scrollable row.
    rows = [_row("F1", "A-1", "done", start=dt.date(2026, 1, 5), end=dt.date(2026, 1, 9))]
    wb, _, weekly = _new_sheets()
    jr.build_weekly_sheet(weekly, rows, REPORT_DATE, {}, feature_names=["F1"])
    assert weekly.freeze_panes == "A6"


# ── Tasks: header row AND column A are both frozen so they stay visible on
# both horizontal and vertical scroll (in Excel and, on upload, Google Sheets
# reading the same OOXML freeze-pane metadata) ──────────────────────────────

def test_tasks_freezes_header_row_and_column_a(tmp_path):
    rows = [_row("F1", "A-1", "done", start=dt.date(2026, 1, 5), end=dt.date(2026, 1, 9))]
    out_path = tmp_path / "report.xlsx"
    jr.build_xlsx(rows, str(out_path), report_date=REPORT_DATE)
    wb = openpyxl.load_workbook(str(out_path))
    assert wb["tasks"].freeze_panes == "B2"


# ── Tasks: Rank column sits between Substream and Task type ────────────────

def test_tasks_rank_column_sits_between_substream_and_task_type(tmp_path):
    row = _row("F1", "A-1", "done", start=dt.date(2026, 1, 5), end=dt.date(2026, 1, 9))
    row["Rank"] = 2
    out_path = tmp_path / "report.xlsx"
    jr.build_xlsx([row], str(out_path), report_date=REPORT_DATE)
    sheet = openpyxl.load_workbook(str(out_path))["tasks"]
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    assert headers.index("Rank") == headers.index("Substream") + 1
    assert headers.index("Task type") == headers.index("Rank") + 1
    rank_col = headers.index("Rank") + 1
    assert sheet.cell(2, rank_col).value == 2


def test_tasks_rank_blank_when_row_has_no_rank_value(tmp_path):
    # A row that never went through assign_rank_numbers() (e.g. an older
    # call site, or a row with no board rank data) must not crash or show a
    # placeholder — just blank, same as "Feature status" already does.
    row = _row("F1", "A-1", "done", start=dt.date(2026, 1, 5), end=dt.date(2026, 1, 9))
    out_path = tmp_path / "report.xlsx"
    jr.build_xlsx([row], str(out_path), report_date=REPORT_DATE)
    sheet = openpyxl.load_workbook(str(out_path))["tasks"]
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    rank_col = headers.index("Rank") + 1
    assert sheet.cell(2, rank_col).value is None


def test_tasks_raw_rank_column_is_written_and_hidden(tmp_path):
    # Raw Jira Rank (LexoRank string) must round-trip through the file
    # itself (not just live in memory for the run that fetched it) so a
    # later --update run, which never refetches an already-done task, can
    # still recompute that task's visible Rank number correctly instead of
    # losing it — see test_jira_report.py's
    # test_update_splice_no_longer_wipes_rank_for_an_unrefreshed_done_task.
    # Kept as a separate "RawRank" column from the visible "Rank" number —
    # it's a real value users don't need to see, hence hidden.
    row = _row("F1", "A-1", "done", start=dt.date(2026, 1, 5), end=dt.date(2026, 1, 9))
    row["_rank"] = "1|i0001:"
    out_path = tmp_path / "report.xlsx"
    jr.build_xlsx([row], str(out_path), report_date=REPORT_DATE)
    sheet = openpyxl.load_workbook(str(out_path))["tasks"]
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    raw_rank_col = headers.index("RawRank") + 1
    assert sheet.cell(2, raw_rank_col).value == "1|i0001:"
    raw_rank_letter = sheet.cell(1, raw_rank_col).column_letter
    assert sheet.column_dimensions[raw_rank_letter].hidden is True


# ── main() end-to-end: resync/update must never leave tasks without Rank ──
#
# These drive the real main() with a mocked jget (the one Jira network choke
# point — see jget()/fetch_all_search()), an on-disk "existing report" xlsx,
# and a throwaway state file. Regression coverage for two real bugs that
# only exist at this level (not visible to any pure-function unit test):
# resync silently skipping already-correctly-placed tasks (never refreshing
# their Rank/title/anything), and resync running a wasteful, unrelated
# report-wide status scan on top of its own epic-scoped fetch.

import json
import urllib.parse


def _issue(key, fields):
    return {"key": key, "fields": fields}


def _make_jget_router(field_list, epics_by_key, epic_search_results, child_issues_by_epic):
    """Fake for jr.jget dispatching on the JQL embedded in the URL — the
    real shape fetch_all_search()/detect_jira_fields() build."""
    calls = []

    def _fake_jget(url, context=None):
        calls.append(url)
        if url.endswith("/field"):
            return field_list
        query = urllib.parse.parse_qs(urllib.parse.urlparse(url).query)
        jql = query.get("jql", [""])[0]
        if jql.startswith("key in ("):
            keys = [k.strip() for k in jql[len("key in ("):-1].split(",")]
            matched = [epics_by_key[k] for k in keys if k in epics_by_key]
            return {"issues": matched, "total": len(matched)}  # total must match len(issues), or fetch_all_search's pagination loop spins forever
        if "issuetype = Epic" in jql:
            return {"issues": epic_search_results, "total": len(epic_search_results)}
        if '"Epic Link" =' in jql:
            issues = []
            for epic_key, epic_issues in child_issues_by_epic.items():
                if f'"Epic Link" = {epic_key}' in jql:
                    issues.extend(epic_issues)
            return {"issues": issues, "total": len(issues)}
        if jql.startswith("issue in ("):
            return {"issues": [], "total": 0}  # the report-wide status scan, if it runs at all
        raise AssertionError(f"unexpected jql in test: {jql!r}")

    return _fake_jget, calls


_FIELD_LIST = [
    {"id": "customfield_10014", "name": "Epic Link"},
    {"id": "customfield_10011", "name": "Epic Name"},
    {"id": "customfield_10100", "name": "Rank"},
]


def _write_state(tmp_path, output_path, **extra):
    state_path = tmp_path / "state.json"
    state_path.write_text(json.dumps({
        "features": [{"keyword": "Checkout Redesign"}],
        "exclude": [],
        "project_keys": [],
        "output": str(output_path),
        "local_only": True,
        "_auto_generated": {},
        **extra,
    }))
    return state_path


def test_update_preserves_settings_keys_it_does_not_manage(tmp_path, monkeypatch):
    # Regression: main()'s end-of-run state write used to construct a brand
    # new dict from scratch instead of updating the loaded one in place,
    # silently dropping any key this script doesn't itself manage
    # (done_statuses, auto_update, or anything roadmap-launcher.py /
    # run-daily-update.py store there) on every single run — including the
    # real daily automation, since it shells out to this exact --update
    # path. Caught live: a real settings file lost a custom done_statuses
    # and its auto_update flag after nothing more than a normal --update.
    monkeypatch.chdir(tmp_path)
    existing_rows = [_row("Checkout Redesign", "TASK-1", "done", start=dt.date(2026, 1, 5), end=dt.date(2026, 1, 9))]
    existing_rows[0]["Epic"] = "EPIC-1"
    output_path = tmp_path / "existing-report.xlsx"
    jr.build_xlsx(existing_rows, str(output_path), report_date=REPORT_DATE)

    epic = _issue("EPIC-1", {"summary": "Checkout Redesign: core", "issuetype": {"name": "Epic"}, "project": {"key": "ABC"}})
    fake_jget, _ = _make_jget_router(
        field_list=_FIELD_LIST, epics_by_key={"EPIC-1": epic},
        epic_search_results=[], child_issues_by_epic={},
    )
    monkeypatch.setattr(jr, "jget", fake_jget)
    monkeypatch.setattr(jr, "jira_headers", lambda: {})

    state_path = _write_state(
        tmp_path, output_path,
        done_statuses=["Shipped", "Verified"],  # deliberately non-default, so this can't pass by accident
        auto_update=False,
    )
    monkeypatch.setattr(sys, "argv", [
        "jira-report.py", "--state", str(state_path), "--output", str(output_path), "--update",
    ])

    try:
        jr.main()
    except SystemExit:
        pass

    written = json.loads(state_path.read_text())
    assert written["done_statuses"] == ["Shipped", "Verified"]
    assert written["auto_update"] is False


def test_resync_refreshes_already_placed_tasks_and_backfills_rank(tmp_path, monkeypatch):
    # Regression: resync's _relabel_or_add() used to `return` early for a
    # task already correctly placed under its feature — skipping the
    # replacement entirely, so its title/status/Rank stayed frozen at
    # whatever they were the last time it genuinely needed relabeling.
    #
    # Defense in depth: main() falls back to a CWD-relative default output
    # path (report/roadmap {year}.xlsx) whenever --output isn't passed on
    # argv exactly right — this happened while writing this test and it
    # silently wrote to the real project's real report file. --output is
    # passed explicitly below to prevent that, but chdir into tmp_path too
    # so even a future mistake here can only ever touch a throwaway path.
    monkeypatch.chdir(tmp_path)
    existing_rows = [
        _row("Checkout Redesign", "TASK-1", "done", start=dt.date(2026, 1, 5), end=dt.date(2026, 1, 9)),
        _row("Checkout Redesign", "TASK-2", "in progress", start=dt.date(2026, 1, 10)),
    ]
    for row in existing_rows:
        row["Epic"] = "EPIC-1"
    existing_rows[0]["Task"] = "old title for TASK-1"
    existing_rows[0]["_rank"] = "1|i0001:"
    existing_rows[1]["_rank"] = "1|i0002:"
    existing_rows = jr.assign_rank_numbers(existing_rows)
    assert existing_rows[0]["Rank"] == 1 and existing_rows[1]["Rank"] == 2  # sanity on the fixture itself

    output_path = tmp_path / "existing-report.xlsx"
    jr.build_xlsx(existing_rows, str(output_path), report_date=REPORT_DATE)

    epic = _issue("EPIC-1", {"summary": "Checkout Redesign: core", "issuetype": {"name": "Epic"}, "project": {"key": "ABC"}})
    fresh_task_1 = _issue("TASK-1", {
        "summary": "new title for TASK-1", "issuetype": {"name": "Task"},
        "status": {"name": "Done"}, "created": "2026-01-05T00:00:00.000-0000",
        "resolutiondate": "2026-01-09T00:00:00.000-0000",
        "customfield_10014": {"key": "EPIC-1"}, "customfield_10100": "1|i0003:",
    })
    fresh_task_2 = _issue("TASK-2", {
        "summary": "task TASK-2", "issuetype": {"name": "Task"},
        "status": {"name": "In Progress"}, "created": "2026-01-10T00:00:00.000-0000",
        "resolutiondate": None,
        "customfield_10014": {"key": "EPIC-1"}, "customfield_10100": "1|i0002:",
    })
    fake_jget, calls = _make_jget_router(
        field_list=_FIELD_LIST,
        epics_by_key={"EPIC-1": epic},
        epic_search_results=[epic],
        child_issues_by_epic={"EPIC-1": [fresh_task_1, fresh_task_2]},
    )
    monkeypatch.setattr(jr, "jget", fake_jget)
    monkeypatch.setattr(jr, "jira_headers", lambda: {})

    state_path = _write_state(tmp_path, output_path)
    monkeypatch.setattr(sys, "argv", [
        # --output MUST be passed explicitly here, not just via the state
        # file's "output" key — main() only honors state's "output" when
        # args.output (the raw CLI flag) is also non-empty; omitting it
        # silently redirects everything to the real default report path
        # (report/roadmap {year}.xlsx in the CWD) instead of this tmp file.
        "jira-report.py", "--state", str(state_path), "--output", str(output_path),
        "--update", "--new-features", "Checkout Redesign",
    ])

    with pytest.raises(SystemExit):
        jr.main()

    result_sheet = openpyxl.load_workbook(str(output_path))["tasks"]
    headers = [result_sheet.cell(1, c).value for c in range(1, result_sheet.max_column + 1)]
    link_col = headers.index("Link") + 1
    task_col = headers.index("Task") + 1
    rank_col = headers.index("Rank") + 1
    rows_by_link = {}
    for r in range(2, result_sheet.max_row + 1):
        link = result_sheet.cell(r, link_col).value
        rows_by_link[link] = (result_sheet.cell(r, task_col).value, result_sheet.cell(r, rank_col).value)

    assert rows_by_link["TASK-1"][0] == "new title for TASK-1"  # the fix: title actually refreshed
    assert rows_by_link["TASK-1"][1] not in (None, "")  # every task must have Rank
    assert rows_by_link["TASK-2"][1] not in (None, "")

    assert not any(u.split("jql=")[-1].startswith("issue%20in%20") for u in calls), (
        "resync must not also run the report-wide not-done status scan"
    )


# ── norm_status: done-status set is configurable, not hardcoded ────────────

def test_norm_status_done_set_is_configurable():
    original = list(jr.DONE_STATUSES)
    try:
        jr.DONE_STATUSES = ["Shipped"]
        jr.DONE_STATUSES_LOWER = {s.lower() for s in jr.DONE_STATUSES}
        assert jr.norm_status("Shipped") == "done"
        assert jr.norm_status("Done") == ""  # no longer in the configured set
    finally:
        jr.DONE_STATUSES = original
        jr.DONE_STATUSES_LOWER = {s.lower() for s in jr.DONE_STATUSES}
