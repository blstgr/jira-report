#!/usr/bin/env python3
import argparse
import datetime as dt
import math
import json
import os
import re
import random
import subprocess
import sys
import time
import threading
import socket
import urllib.parse
import urllib.error
import urllib.request
from copy import copy
from functools import lru_cache
from pathlib import Path

from google_drive_sync import upload_or_update

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Alignment, Border, Side
from openpyxl.worksheet.filters import FilterColumn, Filters
try:
    from PIL import ImageFont
except Exception:
    ImageFont = None


DEFAULT_HOST = os.environ.get("JIRA_HOST", "")
BASE = f"https://{DEFAULT_HOST}/rest/api/2"
TOKEN_SERVICE = os.environ.get("JIRA_KEYCHAIN_SERVICE", "atlassian-dc-mcp")
TOKEN_ACCOUNT = os.environ.get("JIRA_KEYCHAIN_ACCOUNT", "jira-token")
SPEC_PATH = Path(__file__).resolve().parents[1] / "prompts" / "roadmap-spec.json"
OUT_DEFAULT = ""
_SETTINGS_DIR = Path(__file__).resolve().parents[1] / "settings"
_SETTINGS_TEMPLATE = _SETTINGS_DIR / "roadmap-settings.json"
_SETTINGS_LOCAL = _SETTINGS_DIR / "roadmap-settings.local.json"
STATE_DEFAULT = str(_SETTINGS_LOCAL if _SETTINGS_LOCAL.exists() else _SETTINGS_TEMPLATE)
SNAPSHOT_DEFAULT = Path(__file__).resolve().parents[1] / "data" / "raw-jira-snapshot.json"
TODAY = dt.date.today()
NOW = dt.datetime.now()
DEFAULT_ETA_FIELD_IDS = ["customfield_19206", "customfield_19204"]
MAX_EPICS_PER_KEYWORD = 100
EPIC_BATCH_SIZE = 10
REQUEST_PAUSE_SECONDS = 0.25
EPIC_KEY_RE = re.compile(r"^[A-Z][A-Z0-9]+-\d+$")
SYMBOLS = ["◐", "◓", "◑", "◒", "✦", "✧", "⬣", "⬢"]
TERM_WIDTH = 100
DEBUG = False
PROJECT_KEYS = []  # empty = accept all; populated from settings at startup
DONE_STATUSES = ["Done", "QA Prod Done", "In Validation"]  # populated from settings at startup
DONE_STATUSES_LOWER = {s.lower() for s in DONE_STATUSES}


class JiraAuthError(RuntimeError):
    pass


class JiraNetworkError(RuntimeError):
    pass


def spinner_line(symbol, message):
    text = f"{symbol} {message}"
    if len(text) > TERM_WIDTH - 1:
        text = text[: TERM_WIDTH - 4] + "..."
    return f"\r\x1b[2K{text}"


def say_inline(message):
    sys.stdout.write(spinner_line(random.choice(SYMBOLS), message) + "\n")
    sys.stdout.flush()


def say_done(message):
    sys.stdout.write(f"\r\x1b[2K✓ {message}\n")
    sys.stdout.flush()


def say_debug(message):
    if not DEBUG:
        return
    sys.stdout.write(f"• {message}\n")
    sys.stdout.flush()


def run_spinner(message, work_fn):
    stop = threading.Event()

    def animate():
        idx = 0
        while not stop.is_set():
            sys.stdout.write(spinner_line(SYMBOLS[idx % len(SYMBOLS)], message))
            sys.stdout.flush()
            idx += 1
            stop.wait(0.12)

    thread = threading.Thread(target=animate, daemon=True)
    thread.start()
    try:
        return work_fn()
    finally:
        stop.set()
        thread.join(timeout=1)
        sys.stdout.write(f"\r\x1b[2K✓ {message}\n")
        sys.stdout.flush()


def run_progress_spinner(initial_message, work_fn):
    stop = threading.Event()
    state = {"message": initial_message}
    lock = threading.Lock()

    def set_message(message):
        with lock:
            state["message"] = message

    def animate():
        idx = 0
        while not stop.is_set():
            with lock:
                message = state["message"]
            sys.stdout.write(spinner_line(SYMBOLS[idx % len(SYMBOLS)], message))
            sys.stdout.flush()
            idx += 1
            stop.wait(0.12)

    thread = threading.Thread(target=animate, daemon=True)
    thread.start()
    try:
        return work_fn(set_message)
    finally:
        stop.set()
        thread.join(timeout=1)
        sys.stdout.write(f"\r\x1b[2K✓ {state['message']}\n")
        sys.stdout.flush()


def normalize_keyword(value):
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def normalized_contains(text, needle):
    return normalize_keyword(needle) in normalize_keyword(text)


def jql_quote(value):
    return str(value).replace("\\", "\\\\").replace('"', '\\"')


def load_report_spec():
    if SPEC_PATH.exists():
        return json.loads(SPEC_PATH.read_text())
    return {
        "feature_keyword": None,
        "exclude_keywords": ["post release", "post-release"],
    }


def snapshot_path(path_value=None):
    path = Path(path_value or SNAPSHOT_DEFAULT)
    if not path.is_absolute():
        path = (Path(__file__).resolve().parents[1] / path).resolve()
    return path


def write_snapshot(path, payload):
    path = snapshot_path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, default=str))


def read_snapshot(path):
    path = snapshot_path(path)
    if not path.exists():
        return None
    return json.loads(path.read_text())


def current_output_path():
    return Path("report") / f"roadmap {TODAY.year}.xlsx"


FEATURE_PALETTE = [
    "DCEBFA",
    "FBE4D8",
    "E2F2E6",
    "F6E3F3",
    "F9EDC9",
    "DDEBF7",
    "EADCF8",
    "DFF3F0",
    "F8E1E7",
    "E7F0D8",
    "D7E8F7",
    "F7E0D7",
    "DDEFD9",
    "F7DCEB",
    "F8ECCF",
    "D9EAF8",
    "E4D8F7",
    "D6F0EC",
    "F9DDE2",
    "E6F2D8",
    "D0E6F9",
    "F5DED1",
    "D9F0E0",
    "F2DCEF",
    "FBF0D8",
    "D8E4FA",
    "E9D6F7",
    "D4F1EF",
    "F8D9DF",
    "E0F0D4",
    "CCE4F7",
    "F7E1D3",
    "D8F0DA",
    "F3DAF0",
    "F7E7CE",
    "D7E1FA",
    "E7D9F6",
    "D3EFE9",
    "FADBE2",
    "E1F1D7",
    "CFE2F6",
    "F6DDD0",
    "D9F2DB",
    "F0D8F2",
    "F8E9D0",
    "D9E3FA",
    "E6D9F7",
    "D1F0EA",
    "F9DCE0",
    "E3F3D8",
]


def hex_to_rgb(value):
    value = value.strip().lstrip("#")
    return tuple(int(value[i:i + 2], 16) for i in (0, 2, 4))


def readable_text_color(fill_hex):
    r, g, b = hex_to_rgb(fill_hex)
    luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
    return "FF1F1F1F" if luminance > 0.65 else "FFFFFFFF"


def feature_color_map(rows):
    features = []
    for row in rows:
        feature = str(row.get("Feature") or "").strip()
        if feature and feature not in features:
            features.append(feature)
    mapping = {}
    for idx, feature in enumerate(features):
        fill_hex = FEATURE_PALETTE[idx % len(FEATURE_PALETTE)]
        mapping[feature] = {
            "fill": fill_hex,
            "font": readable_text_color(fill_hex),
        }
    return mapping


def annotate_feature_status(rows):
    feature_issue_rows = {}
    for row in rows:
        feature = str(row.get("Feature") or "").strip()
        issue_key = str(row.get("Link") or "").strip()
        if not feature or not issue_key:
            continue
        feature_issue_rows.setdefault(feature, {}).setdefault(issue_key, []).append(row)

    feature_state = {}
    for feature, issue_groups in feature_issue_rows.items():
        state = {"done": 0, "rejected": 0, "in_progress": 0, "blank": 0, "on_hold": 0}
        for issue_rows_list in issue_groups.values():
            current_row = current_issue_row(issue_rows_list)
            status = str(current_row.get("Status") or "").strip()
            if status == "done":
                state["done"] += 1
            elif status == "rejected":
                state["rejected"] += 1
            elif status == "in progress":
                state["in_progress"] += 1
            elif status == "on hold":
                state["on_hold"] += 1
            else:
                state["blank"] += 1
        feature_state[feature] = state

    for row in rows:
        feature = str(row.get("Feature") or "").strip()
        state = feature_state.get(feature, {})
        if not feature:
            feature_status = ""
        elif state.get("in_progress", 0) > 0:
            feature_status = ""
        elif state.get("done", 0) > 0 or state.get("rejected", 0) > 0:
            if state.get("blank", 0) == 0 and state.get("on_hold", 0) == 0:
                feature_status = "done"
            else:
                feature_status = "on hold"
        elif state.get("blank", 0) > 0 or state.get("on_hold", 0) > 0:
            feature_status = "open"
        else:
            feature_status = ""
        row["Feature status"] = feature_status
    return rows


def is_epic_key_keyword(value):
    return bool(EPIC_KEY_RE.match((value or "").strip().upper()))


def all_project_name(value):
    text = (value or "").strip()
    if text.lower().startswith("all ") and len(text.split(None, 1)) == 2:
        return text.split(None, 1)[1].strip()
    return ""


def project_matches_selector(fields, selector):
    project = fields.get("project") or {}
    project_name = (project.get("name") or "").strip().lower()
    project_key = (project.get("key") or "").strip().lower()
    wanted = (selector or "").strip().lower()
    return wanted in {project_name, project_key}


def keyword_match_rank(text, keyword):
    normalized_text = normalize_keyword(text)
    normalized_keyword = normalize_keyword(keyword)
    if not normalized_keyword or not normalized_text:
        return None
    if normalized_text == normalized_keyword:
        return (0, len(keyword))
    if normalized_text.startswith(normalized_keyword):
        return (1, -len(keyword))
    if normalized_keyword in normalized_text:
        return (2, -len(keyword))
    return None


def pick_feature_label(epic_summary, epic_name, include_values, epic_key="", return_source=False):
    best = None
    for include in include_values:
        include_text = (include or "").strip()
        if not include_text:
            continue
        if is_epic_key_keyword(include_text):
            if include_text.upper() == (epic_key or "").strip().upper():
                label = (epic_name or epic_summary or include_text).strip() or include_text
                candidate = (-1, -len(include_text), 0, include_text.lower(), label, include_text)
                if best is None or candidate < best:
                    best = candidate
            continue
        project_name = all_project_name(include_text)
        if project_name:
            candidate = (9, 0, 0, include_text.lower(), include_text, include_text)
            if best is None or candidate < best:
                best = candidate
            continue
        for source_name, source_value in (("epic_name", epic_name), ("summary", epic_summary)):
            rank = keyword_match_rank(source_value or "", include_text)
            if rank is None:
                continue
            candidate = (rank[0], rank[1], 0 if source_name == "epic_name" else 1, include_text.lower(), include_text, include_text)
            if best is None or candidate < best:
                best = candidate
    if best:
        return (best[-2], best[-1]) if return_source else best[-2]
    fallback = include_values[0] if include_values else ""
    return (fallback, fallback) if return_source else fallback


def lookup_expected_rate(expected_tasks_per_week, key):
    target = str(key or "").strip()
    if not target:
        return ""
    for existing_key, value in (expected_tasks_per_week or {}).items():
        if str(existing_key).strip().lower() == target.lower():
            return value
    return ""


def parse_freeform_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value).strip()
    if not text:
        return None
    current_year = dt.date.today().year
    formats = [
        "%b %d, %Y",
        "%d %b %Y",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%m-%d-%Y",
        "%b %d %Y",
        "%B %d, %Y",
        "%d %B %Y",
        "%B %d %Y",
        "%m/%d/%Y",
        "%d/%m/%Y",
        "%b %d",
        "%d %b",
        "%B %d",
        "%d %B",
        "%m-%d",
        "%d-%m",
        "%m/%d",
        "%d/%m",
    ]
    for fmt in formats:
        try:
            parsed = dt.datetime.strptime(text, fmt)
        except Exception:
            continue
        if "%Y" not in fmt:
            parsed = parsed.replace(year=current_year)
        return parsed.date()
    return None


def sanitize_feature_eta_dates(values):
    cleaned = {}
    for key, value in (values or {}).items():
        feature = str(key).strip()
        if not feature:
            continue
        parsed = parse_freeform_date(value)
        if parsed:
            cleaned[feature] = parsed.isoformat()
    return cleaned


def lookup_feature_eta_date(feature_eta_dates, key):
    target = str(key or "").strip()
    if not target:
        return None
    for existing_key, value in (feature_eta_dates or {}).items():
        if str(existing_key).strip().lower() == target.lower():
            try:
                return dt.date.fromisoformat(str(value))
            except Exception:
                return None
    return None


def keychain_token():
    env_token = os.environ.get("JIRA_TOKEN", "").strip()
    if env_token:
        return env_token
    cmd = (
        "security find-generic-password "
        f"-a {sh_quote(TOKEN_ACCOUNT)} "
        f"-s {sh_quote(TOKEN_SERVICE)} "
        "-w"
    )
    last_error = None
    for attempt in range(3):
        try:
            return subprocess.check_output(["/bin/zsh", "-lc", cmd], text=True).strip()
        except subprocess.CalledProcessError as exc:
            last_error = exc
            time.sleep(0.5 * (attempt + 1))
    raise RuntimeError(
        "Unable to read Jira token. Set JIRA_TOKEN in the environment or re-run Jira setup."
    ) from last_error


def sh_quote(value):
    return "'" + str(value).replace("'", "'\"'\"'") + "'"


def jira_headers():
    token = keychain_token()
    return {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
    }


def classify_network_error(exc):
    if isinstance(exc, urllib.error.URLError):
        reason = getattr(exc, "reason", exc)
        reason_text = str(reason)
    else:
        reason_text = str(exc)
    lower = reason_text.lower()
    if any(token in lower for token in ["timed out", "timeout", "connection reset", "broken pipe", "nodename nor servname", "name or service not known", "temporary failure in name resolution", "network is unreachable", "connection refused"]):
        return True, reason_text
    if isinstance(exc, socket.timeout):
        return True, reason_text
    return False, reason_text


def jget(url, context=None):
    last_error = None
    for attempt in range(1):
        try:
            req = urllib.request.Request(url, headers=jira_headers())
            with urllib.request.urlopen(req, timeout=20) as r:
                return json.loads(r.read().decode())
        except urllib.error.HTTPError as exc:
            if exc.code in {401, 403}:
                raise JiraAuthError(f"Jira authentication failed with HTTP {exc.code}") from exc
            last_error = exc
        except Exception as exc:
            last_error = exc
    is_network, reason_text = classify_network_error(last_error)
    if is_network:
        context_text = f" while {context}" if context else ""
        raise JiraNetworkError(
            f"Jira or VPN connection appears to have dropped{context_text}. "
            f"Last error: {reason_text}. Try fewer keywords or rerun after VPN stabilizes."
        ) from last_error
    raise last_error


def fetch_all_search(jql, fields, expand=None, context=None):
    start = 0
    items = []
    while True:
        params = {"jql": jql, "startAt": start, "maxResults": 100, "fields": fields}
        if expand:
            params["expand"] = expand
        search_context = context or f"running Jira search startAt={start}"
        data = jget(f"{BASE}/search?{urllib.parse.urlencode(params)}", context=search_context)
        items.extend(data["issues"])
        start += len(data["issues"])
        if start >= data["total"]:
            break
        time.sleep(REQUEST_PAUSE_SECONDS)
    return items


def parse_date(value):
    if not value:
        return None
    return dt.datetime.strptime(value[:10], "%Y-%m-%d").date()


def iso_week(value):
    return value.isocalendar()[1] if value else ""


def weeks_between(start, end):
    if not start or not end or end < start:
        return 1
    return max(1, ((end - start).days // 7) + 1)


def in_last_n_weeks(value, report_date, weeks=4):
    if not value:
        return False
    start_date = report_date - dt.timedelta(days=(weeks * 7) - 1)
    return start_date <= value <= report_date


def is_integer_number(value):
    return isinstance(value, (int, float)) and float(value).is_integer()


def set_numeric_format(cell, value, positive_sign=False, percent=False):
    if percent:
        cell.number_format = "0%"
        return
    if not isinstance(value, (int, float)):
        return
    if positive_sign:
        cell.number_format = "\\+0;\\-0;0" if is_integer_number(value) else "\\+0.###;\\-0.###;0"
    else:
        cell.number_format = "0" if is_integer_number(value) else "0.###"


def business_days(start, end):
    if not start:
        return ""
    if not end:
        end = TODAY
    days = 0
    cur = start
    while cur <= end:
        if cur.weekday() < 5:
            days += 1
        cur += dt.timedelta(days=1)
    return days


def business_day_dates(start, end):
    if not start:
        return set()
    if not end:
        end = TODAY
    dates = set()
    cur = start
    while cur <= end:
        if cur.weekday() < 5:
            dates.add(cur)
        cur += dt.timedelta(days=1)
    return dates


def work_weeks_for_row(row, report_date):
    start = row.get("Start")
    if not start:
        return set()
    status = row.get("Status") or ""
    end = row.get("End")
    if status == "in progress":
        end = report_date
    elif not end:
        return set()
    weeks = set()
    for day in business_day_dates(start, end):
        weeks.add(day.isocalendar()[1])
    if status == "done" and row.get("End"):
        weeks.add(row["End"].isocalendar()[1])
    elif not weeks:
        weeks.add(start.isocalendar()[1])
    return weeks


def norm_status(value):
    if not value:
        return ""
    s = value.lower()
    if s == "in progress" or s.startswith("in qa") or s in {"code review", "progress done"}:
        return "in progress"
    if s in DONE_STATUSES_LOWER:
        return "done"
    if s == "rejected":
        return "rejected"
    if s in {"qa on hold", "track/blocked/on hold"}:
        return "on hold"
    return ""


def task_type(issue_type):
    return "Task" if issue_type == "Story" else issue_type


def pick_eta(fields, eta_field_ids):
    for key in eta_field_ids:
        value = fields.get(key)
        if value in (None, ""):
            continue
        if isinstance(value, dict):
            value = value.get("value")
        try:
            return float(value)
        except Exception:
            continue
    return ""


def extract_events(issue):
    events = []
    for history in issue.get("changelog", {}).get("histories", []):
        created_raw = history["created"]
        when = parse_date(created_raw)
        # Keep the full timestamp string as a secondary sort key so that
        # same-day events are ordered chronologically. This prevents a
        # "done" entry that Jira happens to list before an "on hold" entry
        # (within the same calendar day) from being processed in the wrong
        # order, which would otherwise produce a spurious on-hold row plus
        # a separate done row instead of a single merged done row.
        sort_key = (when, created_raw or "")
        for item in history.get("items", []):
            if item.get("field") == "status":
                events.append((sort_key, when, item.get("fromString") or "", item.get("toString") or ""))
    events.sort(key=lambda x: x[0])
    return [(when, frm, to) for _key, when, frm, to in events]


def detect_eta_field_ids():
    try:
        fields = jget(f"{BASE}/field")
    except Exception:
        return list(DEFAULT_ETA_FIELD_IDS)

    scored = []
    for field in fields:
        field_id = field.get("id")
        name = (field.get("name") or "").strip()
        if not field_id or not name:
            continue
        lowered = name.lower()
        score = None
        if lowered == "eta":
            score = 0
        elif lowered.startswith("eta "):
            score = 1
        elif " eta" in lowered or "eta" in lowered:
            score = 2
        if score is not None:
            scored.append((score, name, field_id))

    ordered = [field_id for _score, _name, field_id in sorted(scored)]
    for fallback in DEFAULT_ETA_FIELD_IDS:
        if fallback not in ordered:
            ordered.append(fallback)
    return ordered


def detect_epic_link_field_id():
    try:
        fields = jget(f"{BASE}/field")
    except Exception:
        return None

    for field in fields:
        field_id = field.get("id")
        name = (field.get("name") or "").strip()
        if field_id and name.lower() == "epic link":
            return field_id
    return None


def detect_epic_name_field_ids():
    try:
        fields = jget(f"{BASE}/field")
    except Exception:
        return ["customfield_10011"]

    scored = []
    for field in fields:
        field_id = field.get("id")
        name = (field.get("name") or "").strip()
        if not field_id or not name:
            continue
        lowered = name.lower()
        score = None
        if lowered == "epic name":
            score = 0
        elif "epic name" in lowered:
            score = 1
        if score is not None:
            scored.append((score, name, field_id))

    ordered = [field_id for _score, _name, field_id in sorted(scored)]
    if "customfield_10011" not in ordered:
        ordered.append("customfield_10011")
    return ordered


def cached_field_ids(state, key, fallback):
    # Check _cache block first, fall back to top-level for old files
    value = state.get("_auto_generated", {}).get(key) or state.get(key)
    if isinstance(value, list) and value:
        return value
    return list(fallback)


def row_belongs_in_current_report_year(row, report_year):
    start = row.get("Start")
    end = row.get("End")
    lookback_year = report_year - 1
    if start and start.year < lookback_year and (not end or end.year < lookback_year):
        return False
    if end and end.year < lookback_year and not start:
        return False
    return True


def rows_have_report_year_activity(issue_rows, report_year):
    return any(
        (row.get("Start") and row["Start"].year == report_year)
        or (row.get("End") and row["End"].year == report_year)
        for row in issue_rows
    )


def rows_have_feature_report_activity(feature_rows, report_year):
    return any(
        (row.get("Start") and row["Start"].year == report_year)
        or (row.get("End") and row["End"].year == report_year)
        for row in feature_rows
    )


def get_epic_name(fields, epic_name_field_ids):
    for field_id in epic_name_field_ids:
        value = fields.get(field_id)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def epic_matches_keyword(epic, keyword, epic_name_field_ids):
    fields = epic.get("fields", {})
    summary = fields.get("summary", "") or ""
    epic_name = get_epic_name(fields, epic_name_field_ids)
    epic_key = (epic.get("key") or "").strip().upper()
    keyword_text = (keyword or "").strip()
    if is_epic_key_keyword(keyword_text):
        return epic_key == keyword_text.upper()
    project_name = all_project_name(keyword_text)
    if project_name:
        return project_matches_selector(fields, project_name)
    return normalized_contains(summary, keyword_text) or normalized_contains(epic_name, keyword_text)


def epic_matches_any_keyword(epic, keywords, epic_name_field_ids):
    return any(epic_matches_keyword(epic, keyword, epic_name_field_ids) for keyword in keywords)


def substream_name(epic_name, epic_summary, feature):
    epic_name_clean = (epic_name or "").strip()
    epic_summary_clean = (epic_summary or "").strip()
    if epic_name_clean and epic_summary_clean and epic_name_clean != epic_summary_clean:
        source = epic_summary_clean
    else:
        source = epic_name_clean or epic_summary_clean
    parts = feature if isinstance(feature, list) else [feature]
    value = source
    for part in parts:
        keyword = (part or "").strip()
        if not keyword:
            continue
        escaped = re.escape(keyword)
        pattern = r"^\s*" + escaped + r"(?:\s*[:\-]\s*|\s+)?"
        value = re.sub(pattern, "", value, flags=re.I).strip()
    return value


def slugify(value):
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def epic_list_link(epic_keys):
    if not epic_keys:
        return ""
    quoted = ",".join(f'"{key}"' for key in sorted(epic_keys))
    jql = f"key in ({quoted})"
    return f"https://{DEFAULT_HOST}/issues/?jql={urllib.parse.quote(jql)}"


def pixels_to_excel_width(pixels):
    if not pixels:
        return 0
    # Approximate conversion from Google Sheets / Excel pixel sizing
    # to openpyxl's character-based column width units.
    return round(max((float(pixels) - 5) / 7, 1), 2)


@lru_cache(maxsize=1)
def excel_arial_font():
    if ImageFont is None:
        return None
    candidates = [
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/Library/Fonts/Arial.ttf",
        "/Library/Fonts/Arial Unicode.ttf",
    ]
    for path in candidates:
        if Path(path).exists():
            return ImageFont.truetype(path, 10)
    return ImageFont.load_default()


def cell_display_text(cell):
    value = cell.value
    if value in (None, ""):
        return ""
    if isinstance(value, dt.datetime):
        value = value.date()
    if isinstance(value, dt.date):
        return value.strftime("%d.%b.%y")
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value)


def _text_width_px(text):
    font = excel_arial_font()
    if not text:
        return 0
    if font is None:
        return len(text) * 7
    try:
        return font.getlength(text)
    except Exception:
        bbox = font.getbbox(text)
        return max(0, bbox[2] - bbox[0])


def _wrap_lines_for_width(text, max_px):
    if not text:
        return 1
    paragraphs = str(text).split("\n")
    lines = 0
    for paragraph in paragraphs:
        paragraph = paragraph.strip()
        if not paragraph:
            lines += 1
            continue
        words = paragraph.split()
        if not words:
            lines += 1
            continue
        current = ""
        for word in words:
            candidate = word if not current else f"{current} {word}"
            if _text_width_px(candidate) <= max_px:
                current = candidate
                continue
            if current:
                lines += 1
            if _text_width_px(word) <= max_px:
                current = word
                continue
            chunk = ""
            for char in word:
                candidate = chunk + char
                if _text_width_px(candidate) <= max_px:
                    chunk = candidate
                else:
                    if chunk:
                        lines += 1
                    chunk = char
            current = chunk
        if current:
            lines += 1
    return max(lines, 1)


def estimate_row_height(sheet, row_idx, *, min_height=18, line_height=15, scale_on_wrap=False):
    max_lines = 1
    for cell in sheet[row_idx]:
        if isinstance(cell.value, str) and cell.value.startswith("="):
            # Can't measure the rendered result, but CHAR(10) in the formula
            # means at least one newline in the output — count occurrences.
            newlines = cell.value.count("CHAR(10)")
            if newlines:
                max_lines = max(max_lines, newlines + 1)
            continue
        text = cell_display_text(cell)
        if not text:
            continue
        if "\n" in text:
            max_lines = max(max_lines, text.count("\n") + 1)
            continue
        column_width = sheet.column_dimensions[cell.column_letter].width
        if column_width in (None, 0):
            column_width = 8.43
        max_px = float(column_width) * 7 + 5
        lines = _wrap_lines_for_width(text, max_px)
        max_lines = max(max_lines, lines)
    if scale_on_wrap:
        return max(min_height, min_height * max_lines)
    return max(min_height, max_lines * line_height)


def auto_fit_row_heights(sheet, start_row, end_row, *, min_height=18, line_height=15, scale_on_wrap=False):
    for row_idx in range(start_row, end_row + 1):
        if row_idx in sheet.row_dimensions and sheet.row_dimensions[row_idx].hidden:
            continue
        sheet.row_dimensions[row_idx].height = estimate_row_height(
            sheet,
            row_idx,
            min_height=min_height,
            line_height=line_height,
            scale_on_wrap=scale_on_wrap,
        )


def make_project_relative(path_value):
    if not path_value:
        return None
    path = Path(path_value).expanduser()
    try:
        resolved = path.resolve()
    except Exception:
        resolved = path
    project_root = Path(__file__).resolve().parents[1]
    try:
        return str(resolved.relative_to(project_root))
    except Exception:
        return str(path_value)


def resolve_project_path(path_value):
    if not path_value:
        return None
    path = Path(path_value).expanduser()
    if not path.is_absolute():
        path = (Path(__file__).resolve().parents[1] / path).resolve()
    return str(path)


def issue_rows(issue, feature, epic_summary, epic_name, eta_field_ids, epic_key=None):
    fields = issue["fields"]
    issue_key = str(issue.get("key") or "").strip().upper()
    if PROJECT_KEYS and not any(issue_key.startswith(f"{pk}-") for pk in PROJECT_KEYS):
        say_debug(
            f"DEBUG issue skipped by project filter: key={issue.get('key')} "
            f"summary='{(fields.get('summary') or '')[:80]}'"
        )
        return []
    summary = fields.get("summary", "")
    issue_type = fields.get("issuetype", {}).get("name", "")
    status_now = norm_status(fields.get("status", {}).get("name", ""))
    eta = pick_eta(fields, eta_field_ids)
    resolution_date = parse_date(fields.get("resolutiondate"))
    created_date = parse_date(fields.get("created"))
    events = extract_events(issue)

    first_in_progress = None
    first_in_qa = None
    active_start = None
    last_status = None
    active_work_dates = set()
    rows = []
    row_sequence = 0
    pending_terminal_idx = None
    last_hold_idx = None
    last_closed_active_start = None
    resumed_after_hold = False
    resumed_after_hold_had_work = False
    last_open_transition = None

    def push(status, start, end, cumulative_days=None):
        nonlocal row_sequence
        rows.append(
            {
                "status": status,
                "start": start,
                "end": end,
                "cumulative_days": cumulative_days,
                "_seq": row_sequence,
            }
        )
        row_sequence += 1

    def clear_pending_terminal():
        nonlocal pending_terminal_idx, last_hold_idx
        if pending_terminal_idx is not None and pending_terminal_idx == len(rows) - 1:
            rows.pop()
            if last_hold_idx is not None and last_hold_idx >= len(rows):
                last_hold_idx = None
        pending_terminal_idx = None

    for idx, (when, _frm, to) in enumerate(events):
        lowered_to = to.lower()
        next_status = norm_status(to)
        same_day_terminal_ahead = False
        for later_when, _later_from, later_to in events[idx + 1 :]:
            if later_when != when:
                break
            if norm_status(later_to) in {"done", "rejected", "in progress"}:
                same_day_terminal_ahead = True
                break
        if lowered_to.startswith("in qa") and first_in_qa is None:
            first_in_qa = when
        if next_status == "":
            if pending_terminal_idx is not None and pending_terminal_idx == len(rows) - 1:
                converted_start = rows[pending_terminal_idx]["start"]
                converted_end = when
                if converted_start is None:
                    rows.pop()
                    if last_hold_idx is not None and last_hold_idx >= len(rows):
                        last_hold_idx = None
                else:
                    rows[pending_terminal_idx]["status"] = "on hold"
                    rows[pending_terminal_idx]["end"] = converted_end
                    old_hold_idx = last_hold_idx
                    if (
                        old_hold_idx is not None
                        and old_hold_idx != pending_terminal_idx
                        and rows[old_hold_idx]["end"] is not None
                        and converted_start <= rows[old_hold_idx]["end"] + dt.timedelta(days=1)
                    ):
                        prev_end = rows[old_hold_idx]["end"]
                        if converted_end and (prev_end is None or converted_end > prev_end):
                            rows[old_hold_idx]["end"] = converted_end
                        rows.pop()
                        last_hold_idx = old_hold_idx
                    else:
                        last_hold_idx = pending_terminal_idx
                pending_terminal_idx = None
            last_open_transition = when
        if next_status == last_status:
            continue
        if next_status == "in progress":
            if pending_terminal_idx is not None:
                clear_pending_terminal()
            if first_in_progress is None:
                first_in_progress = when
            active_start = when
            resumed_after_hold = last_status == "on hold"
            resumed_after_hold_had_work = False
        elif next_status == "on hold":
            if pending_terminal_idx is not None:
                clear_pending_terminal()
            if same_day_terminal_ahead:
                continue
            start = active_start or last_closed_active_start or first_in_progress or first_in_qa
            if start is not None:
                last_closed_active_start = start
                segment_dates = business_day_dates(start, when)
                new_dates = segment_dates - active_work_dates
                if new_dates:
                    active_work_dates |= segment_dates
                    resumed_after_hold_had_work = True
                previous_hold_end = rows[last_hold_idx]["end"] if last_hold_idx is not None else None
                if (
                    last_hold_idx is not None
                    and previous_hold_end is not None
                    and start is not None
                    and start <= previous_hold_end + dt.timedelta(days=1)
                ):
                    if when and when > previous_hold_end:
                        rows[last_hold_idx]["end"] = when
                    rows[last_hold_idx]["cumulative_days"] = len(active_work_dates) if active_work_dates else ""
                elif last_hold_idx is not None and resumed_after_hold and not resumed_after_hold_had_work:
                    existing_end = rows[last_hold_idx]["end"]
                    if existing_end is None or (when and when > existing_end):
                        rows[last_hold_idx]["end"] = when
                    rows[last_hold_idx]["cumulative_days"] = len(active_work_dates) if active_work_dates else ""
                else:
                    push("on hold", start, when, len(active_work_dates) if active_work_dates else "")
                    last_hold_idx = len(rows) - 1
            active_start = None
            resumed_after_hold = False
            resumed_after_hold_had_work = False
        elif next_status in {"done", "rejected"}:
            if pending_terminal_idx is not None:
                clear_pending_terminal()
            end = when or resolution_date
            start = active_start or (last_closed_active_start if last_status == "on hold" else None) or first_in_progress or first_in_qa
            if start is None and next_status == "done":
                start = end
            if start and end and start > end:
                start = end
            segment_dates = business_day_dates(start, end)
            cumulative_days = ""
            if segment_dates:
                new_dates = segment_dates - active_work_dates
                if new_dates:
                    active_work_dates |= segment_dates
                    resumed_after_hold_had_work = True
                cumulative_days = len(active_work_dates)
            if last_status == "on hold" and active_start is None and last_hold_idx is not None:
                rows[last_hold_idx]["status"] = next_status
                rows[last_hold_idx]["start"] = start
                rows[last_hold_idx]["end"] = end
                rows[last_hold_idx]["cumulative_days"] = cumulative_days
                pending_terminal_idx = last_hold_idx
            else:
                push(next_status, start, end, cumulative_days)
                pending_terminal_idx = len(rows) - 1
            if active_start:
                last_closed_active_start = active_start
            active_start = None
            resumed_after_hold = False
            resumed_after_hold_had_work = False
        last_status = next_status

    if status_now == "in progress":
        if not rows or rows[-1]["status"] != "in progress":
            start = active_start or first_in_progress or first_in_qa
            segment_days = business_day_dates(start, None)
            cumulative_days = ""
            if segment_days:
                cumulative_days = len(active_work_dates | segment_days)
            push("in progress", start, None, cumulative_days)
    elif status_now == "on hold":
        if not rows:
            start = active_start or first_in_progress or first_in_qa
            push("on hold", start, None, len(active_work_dates) if active_work_dates else "")
    elif status_now in {"done", "rejected"} and not any(row["status"] == status_now for row in rows):
        start = first_in_progress or first_in_qa
        if not start and status_now == "done":
            start = resolution_date
        segment_days = business_day_dates(start, resolution_date)
        cumulative_days = ""
        if segment_days:
            active_work_dates |= segment_days
            cumulative_days = len(active_work_dates)
        if last_status == "on hold" and active_start is None and last_hold_idx is not None:
            rows[last_hold_idx]["status"] = status_now
            rows[last_hold_idx]["start"] = start
            rows[last_hold_idx]["end"] = resolution_date
            rows[last_hold_idx]["cumulative_days"] = cumulative_days
        else:
            push(status_now, start, resolution_date, cumulative_days)
    elif status_now == "":
        if not rows or rows[-1]["status"] != "":
            push("", None, None)

    result = []
    substream = substream_name(epic_name, epic_summary, feature)
    if not rows:
        say_debug(
            f"DEBUG no lifecycle rows: key={issue.get('key')} "
            f"status_now={status_now} summary='{summary[:80]}'"
        )
    for row in rows:
        start = row["start"]
        end = row["end"]
        if row["status"] == "done" and not end:
            end = resolution_date
        days = business_days(start, end)
        if row["status"] in {"done", "in progress"} and row.get("cumulative_days") not in ("", None):
            days = row["cumulative_days"]
        delta = ""
        if row["status"] == "done" and eta != "" and days != "":
            delta = days - eta
        result.append(
            {
                "Feature": feature,
                "Epic": epic_key or "",
                "Epic Summary": epic_summary or "",
                "Substream": substream,
                "Task type": task_type(issue_type),
                "Task": summary,
                "Status": row["status"],
                "Start": start,
                "End": end,
                "Done week": iso_week(end),
                "ETA": eta,
                "Days in Work": days,
                "Delta ETA": delta,
                "Link": issue["key"],
                "Created date": created_date,
                "Created week": iso_week(created_date),
                "_seq": row.get("_seq", 0),
            }
        )
    if result:
        say_debug(
            f"DEBUG issue rows created: key={issue.get('key')} "
            f"status_now={status_now} rows={len(result)}"
        )
    return result


def build_xlsx(rows, out_path, expected_tasks_per_week=None, feature_eta_dates=None, report_date=None):
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    expected_tasks_per_week = expected_tasks_per_week or {}
    report_date = report_date or TODAY
    headers = [
        "Feature",
        "Substream",
        "Task type",
        "Task",
        "Status",
        "Start",
        "End",
        "Done week",
        "ETA",
        "Days in Work",
        "Delta ETA",
        "Link",
        "Created date",
        "Created week",
        "Feature status",
    ]
    week_headers = list(range(1, 53))
    fixed_widths = {
        "Feature": pixels_to_excel_width(137),
        "Substream": pixels_to_excel_width(137),
        "Task": pixels_to_excel_width(517),
    }

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    weekly_sheet = workbook.create_sheet("weekly")
    sheet = workbook.create_sheet("tasks")
    sheet.freeze_panes = "A2"

    base_font = Font(name="Arial", size=10)
    header_font = Font(name="Arial", size=10, bold=True)
    in_progress_font = Font(name="Arial", size=10, italic=True)
    link_font = Font(name="Arial", size=10, color="FF0563C1")
    rejected_font = Font(name="Arial", size=10, color="FF808080")
    overdue_font = Font(name="Arial", size=10, color="FFFFFFFF")
    left_middle_alignment = Alignment(horizontal="left", vertical="center")
    feature_styles = feature_color_map(rows)

    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFFF59D")
    gray_fill = PatternFill(fill_type="solid", fgColor="FFE0E0E0")
    red_fill = PatternFill(fill_type="solid", fgColor="FFC62828")
    week_fill = PatternFill(fill_type="solid", fgColor="FF2EC67E")
    week_header_fill = PatternFill(fill_type="solid", fgColor="FFFFF59D")

    all_headers = headers + ["Epic"] + week_headers
    widths = {header: len(str(header)) for header in all_headers}
    current_week = report_date.isocalendar()[1]
    for col_idx, header in enumerate(all_headers, 1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="top")
        if header in week_headers and header == current_week:
            cell.fill = week_header_fill

    for row_idx, row in enumerate(rows, 2):
        active_weeks = work_weeks_for_row(row, report_date)
        for col_idx, header in enumerate(headers, 1):
            value = row.get(header, "")
            if header == "Delta ETA" and (row["Status"] != "done" or row["ETA"] == "" or row["Delta ETA"] == ""):
                value = None
            elif header == "Link":
                value = row["Link"]
            elif value == "":
                value = None

            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = base_font
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            if header == "Feature":
                feature_key = str(value or "").strip()
                feature_style = feature_styles.get(feature_key)
                if feature_style:
                    cell.fill = PatternFill(fill_type="solid", fgColor=f"FF{feature_style['fill']}")
                    cell.font = Font(name="Arial", size=10, color=feature_style["font"])

            if header in {"Start", "End"}:
                cell.number_format = "dd.mmm.yy"
            elif header == "Created date":
                cell.number_format = "dd.mmm.yy"
            elif header in {"Done week", "Created week"}:
                cell.number_format = "0"
            elif header in {"ETA", "Days in Work"}:
                set_numeric_format(cell, value)
            elif header == "Delta ETA":
                set_numeric_format(cell, value, positive_sign=True)

            if header == "Link" and value:
                cell.hyperlink = f"https://{DEFAULT_HOST}/browse/{value}"
                cell.font = link_font

            if row["Status"] == "in progress" and header in {"Task", "Status", "Start"}:
                cell.font = in_progress_font
                cell.fill = yellow_fill

            if row["Status"] == "rejected":
                cell.font = rejected_font
                cell.fill = gray_fill

            if header == "Task" and isinstance(row["Delta ETA"], (int, float)) and row["Delta ETA"] > 1:
                cell.font = overdue_font
                cell.fill = red_fill

            display_value = value
            if isinstance(display_value, dt.date):
                display_value = display_value.strftime("%Y-%m-%d")
            widths[header] = max(widths[header], len("" if display_value is None else str(display_value)))

        epic_col_idx = len(headers) + 1
        epic_cell = sheet.cell(row=row_idx, column=epic_col_idx, value=row.get("Epic") or None)
        epic_cell.font = base_font
        epic_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
        widths["Epic"] = max(widths.get("Epic", 0), len(str(epic_cell.value or "")))
        epic_keys = {k.strip() for k in str(row.get("Epic") or "").split(",") if k.strip()}
        if epic_keys:
            epic_cell.hyperlink = epic_list_link(epic_keys)
            epic_cell.font = link_font

        for offset, week in enumerate(week_headers, len(headers) + 2):
            cell = sheet.cell(row=row_idx, column=offset, value=None)
            cell.font = base_font
            cell.alignment = Alignment(horizontal="left", vertical="top")
            if week in active_weeks:
                cell.fill = week_fill
        if str(row.get("Feature status") or "").strip():
            sheet.row_dimensions[row_idx].hidden = True
    sheet.auto_filter.ref = f"A1:{sheet.cell(row=1, column=len(all_headers)).column_letter}{len(rows) + 1}"
    feature_status_col_idx = all_headers.index("Feature status")
    sheet.auto_filter.filterColumn = []
    sheet.auto_filter.filterColumn.append(FilterColumn(colId=feature_status_col_idx, filters=Filters(blank=True)))
    for col_idx, header in enumerate(all_headers, 1):
        if header == "Epic":
            width = widths[header] + 1
        elif header in week_headers:
            width = pixels_to_excel_width(42)
        else:
            width = fixed_widths.get(header, max(widths[header] + 2, 10))
        sheet.column_dimensions[sheet.cell(row=1, column=col_idx).column_letter].width = width

    auto_fit_row_heights(sheet, 1, len(rows) + 1, min_height=14, line_height=14)

    build_summary_sheet(summary_sheet, rows, expected_tasks_per_week, feature_eta_dates, report_date, feature_styles)
    build_weekly_sheet(
        weekly_sheet,
        rows,
        report_date,
        feature_styles,
        expected_tasks_per_week=expected_tasks_per_week,
        feature_eta_dates=feature_eta_dates,
        feature_names=sorted({str(r.get("Feature") or "").strip() for r in rows if str(r.get("Feature") or "").strip()}, key=str.lower),
    )

    workbook.save(out_path)


def current_status_sort_value(status):
    order = {
        "done": 4,
        "rejected": 3,
        "in progress": 2,
        "on hold": 1,
        "": 0,
        None: 0,
    }
    return order.get(status, 0)


def row_effective_date(row):
    # Rows without an End date are the current/active state — sort them highest.
    end = row.get("End")
    if end:
        return end.date() if isinstance(end, dt.datetime) else end
    return dt.date.max


def feature_issue_groups(rows):
    grouped = {}
    for row in rows:
        feature = str(row.get("Feature") or "").strip()
        issue_key = str(row.get("Link") or "").strip()
        if not feature or not issue_key:
            continue
        grouped.setdefault(feature, {}).setdefault(issue_key, []).append(row)
    return grouped


def current_issue_row(issue_rows_list):
    if not issue_rows_list:
        return None
    return max(
        issue_rows_list,
        key=lambda item: (row_effective_date(item), item.get("_seq", 0)),
    )


def issue_display_link(issue_key):
    issue_key = str(issue_key or "").strip().upper()
    return f"https://{DEFAULT_HOST}/browse/{issue_key}" if issue_key else ""


def humanize_status(state_rows):
    statuses = {str(row.get("Status") or "").strip() for row in state_rows}
    if "in progress" in statuses:
        return "in progress"
    if statuses <= {""}:
        return "open"
    if ("done" in statuses or "rejected" in statuses) and ("" in statuses or "on hold" in statuses):
        return "on hold"
    if statuses <= {"done", "rejected"}:
        return "done"
    return "open"


def blocker_summary_for_row(row):
    task = str(row.get("Task") or "").strip()
    status = str(row.get("Status") or "").strip()
    parts = []
    if status == "on hold":
        parts.append("Work is paused")
    elif status == "in progress":
        parts.append("Work is active")
    elif status == "done":
        parts.append("Work is complete")
    else:
        parts.append("Status is open")
    if task:
        parts.append(f"for '{task}'")
    return " ".join(parts)


def build_weekly_sheet(sheet, rows, report_date, feature_styles, expected_tasks_per_week=None, feature_eta_dates=None, feature_names=None):  # noqa: C901
    base_font = Font(name="Arial", size=10)
    header_font = Font(name="Arial", size=10, bold=True)
    italic_font = Font(name="Arial", size=10, italic=True)
    link_font = Font(name="Arial", size=10, color="FF0563C1")
    top_wrap_alignment = Alignment(vertical="top", wrap_text=True)
    left_top_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    light_green_fill = PatternFill(fill_type="solid", fgColor="FFE2F2D9")
    light_red_fill = PatternFill(fill_type="solid", fgColor="FFFFC7CE")
    subheader_fill = PatternFill(fill_type="solid", fgColor="FFF2F2F2")
    medium_side = Side(style="thin", color="FF0072C6")

    grouped = feature_issue_groups(rows)
    features = sorted(grouped.keys(), key=str.lower)
    all_feature_names = list(feature_names or features)

    # ── synthetic "All" entry — merges every feature's issues into one bucket ──
    ALL_FEATURES_LABEL = "All"
    if ALL_FEATURES_LABEL not in grouped:
        all_issues_group = {}
        for feature_group_src in grouped.values():
            all_issues_group.update(feature_group_src)
        if all_issues_group:
            grouped[ALL_FEATURES_LABEL] = all_issues_group
            if ALL_FEATURES_LABEL not in all_feature_names:
                all_feature_names = [ALL_FEATURES_LABEL] + all_feature_names

    # ── A1: label, A2: feature dropdown ──────────────────────────────────────
    sheet["A1"] = "Feature name"
    sheet["A1"].font = header_font
    sheet["A1"].alignment = left_top_alignment
    sheet["A2"] = ALL_FEATURES_LABEL if ALL_FEATURES_LABEL in grouped else (features[0] if features else "")
    sheet["A2"].font = base_font
    sheet["A2"].alignment = left_top_alignment

    try:
        from openpyxl.worksheet.datavalidation import DataValidation
        if all_feature_names:
            helper = sheet.parent["_weekly_lists"] if "_weekly_lists" in sheet.parent.sheetnames else sheet.parent.create_sheet("_weekly_lists")
            helper.sheet_state = "veryHidden"
            helper.delete_rows(1, helper.max_row)
            for idx, name in enumerate(all_feature_names, 1):
                helper.cell(row=idx, column=1, value=name)
            dv = DataValidation(type="list", formula1=f"'_weekly_lists'!$A$1:$A${len(all_feature_names)}", allow_blank=True)
            sheet.add_data_validation(dv)
            dv.add(sheet["A2"])
    except Exception:
        pass

    # ── _epic_data hidden sheet ───────────────────────────────────────────────
    # Cols: A=Feature, B=EpicKey, C=EpicName, D=Blank, E=InProgress, F=OnHold,
    #        G=Done, H=Rejected, I=Total, J=LookupKey (Feature|seq)
    edata_name = "_epic_data"
    edata = sheet.parent[edata_name] if edata_name in sheet.parent.sheetnames else sheet.parent.create_sheet(edata_name)
    edata.sheet_state = "hidden"
    edata.delete_rows(1, edata.max_row)
    for ci, h in enumerate(["Feature","EpicKey","EpicName","Blank","InProgress","OnHold","Done","Rejected","Total","LookupKey"], 1):
        edata.cell(row=1, column=ci, value=h)

    # ── _feature_req hidden sheet ─────────────────────────────────────────────
    # Cols: A=Feature, B=RequiredPerWeek
    freq_name = "_feature_req"
    fdata = sheet.parent[freq_name] if freq_name in sheet.parent.sheetnames else sheet.parent.create_sheet(freq_name)
    fdata.sheet_state = "hidden"
    fdata.delete_rows(1, fdata.max_row)
    fdata.cell(row=1, column=1, value="Feature")
    fdata.cell(row=1, column=2, value="RequiredPerWeek")

    # ── _weekly_data hidden sheet ─────────────────────────────────────────────
    # Cols: A=Feature, B=WeekNo, C=WeekStart(date), D=TasksDone, E=TasksRejected,
    #        F=BugsDone, G=BugsRejected, H=Total, I=Created, J=LookupKey
    wdata_name = "_weekly_data"
    wdata = sheet.parent[wdata_name] if wdata_name in sheet.parent.sheetnames else sheet.parent.create_sheet(wdata_name)
    wdata.sheet_state = "hidden"
    wdata.delete_rows(1, wdata.max_row)
    for ci, h in enumerate(["Feature","WeekNo","WeekStart","TasksDone","TasksRejected","BugsDone","BugsRejected","Total","Created","LookupKey"], 1):
        wdata.cell(row=1, column=ci, value=h)

    ed_row = 2
    fr_row = 2
    wd_row = 2
    max_feature_epics = 0
    max_feature_weeks = 0

    for feature in all_feature_names:
        feature_group = grouped.get(feature, {})
        if not feature_group:
            continue

        # ── per-epic status counts ────────────────────────────────────────────
        epic_stats: dict = {}
        remaining_count = 0
        by_week: dict = {}
        created_by_week: dict = {}

        for issue_key, issue_rows_list in feature_group.items():
            current_row = current_issue_row(issue_rows_list)
            status = str(current_row.get("Status") or "").strip()
            epic_key = str(current_row.get("Epic") or "").strip() or "(no epic)"
            # Prefer the epic's real Jira summary text; fall back to the derived
            # Substream, then to the feature label, for epics from before this field
            # existed or where the summary was never captured.
            epic_summary_text = next(
                (str(r.get("Epic Summary") or "").strip() for r in issue_rows_list if str(r.get("Epic Summary") or "").strip()),
                "",
            )
            substream = next(
                (str(r.get("Substream") or "").strip() for r in issue_rows_list if str(r.get("Substream") or "").strip()),
                "",
            )
            # Fall back to the issue's own feature, not the loop's `feature` — for the
            # synthetic "All" bucket, `feature` is literally "All" and would otherwise
            # overwrite every substream-less epic's name with "All".
            own_feature = str(current_row.get("Feature") or "").strip() or feature
            epic_display_name = epic_summary_text or substream or own_feature

            if epic_key not in epic_stats:
                epic_stats[epic_key] = {"name": epic_display_name, "blank": 0, "in_progress": 0, "on_hold": 0, "done": 0, "rejected": 0}
            elif not epic_stats[epic_key]["name"] and epic_display_name:
                epic_stats[epic_key]["name"] = epic_display_name

            if status == "done":
                epic_stats[epic_key]["done"] += 1
            elif status == "rejected":
                epic_stats[epic_key]["rejected"] += 1
            elif status == "in progress":
                epic_stats[epic_key]["in_progress"] += 1
            elif status == "on hold":
                epic_stats[epic_key]["on_hold"] += 1
            else:
                epic_stats[epic_key]["blank"] += 1

            if status not in {"done", "rejected"}:
                remaining_count += 1

            # ── created_by_week ───────────────────────────────────────────────
            # Keyed by (year, week) — data can span multiple years (old carried-over
            # epics, etc.), and a bare week number would conflate e.g. week 52 of last
            # year with week 52 of this year.
            cw = current_row.get("Created week")
            created_date = current_row.get("Created date")
            if cw not in (None, ""):
                try:
                    created_week_no = int(cw)
                    created_year = created_date.isocalendar()[0] if isinstance(created_date, dt.date) else report_date.isocalendar()[0]
                    created_key = (created_year, created_week_no)
                    created_by_week[created_key] = created_by_week.get(created_key, 0) + 1
                except (ValueError, TypeError):
                    pass

            # ── by_week (done/rejected only) ──────────────────────────────────
            if status not in {"done", "rejected"}:
                continue
            dw = current_row.get("Done week")
            end_date = current_row.get("End")
            if dw in (None, ""):
                continue
            try:
                week_no = int(dw)
            except (ValueError, TypeError):
                continue
            year = end_date.isocalendar()[0] if isinstance(end_date, dt.date) else report_date.isocalendar()[0]
            week_key = (year, week_no)
            bucket = by_week.setdefault(week_key, {"tasks_done": 0, "tasks_rejected": 0, "bugs_done": 0, "bugs_rejected": 0})
            if str(current_row.get("Task type") or "").strip().lower() == "bug":
                bucket["bugs_done" if status == "done" else "bugs_rejected"] += 1
            else:
                bucket["tasks_done" if status == "done" else "tasks_rejected"] += 1

        # ── required per week for ETA ─────────────────────────────────────────
        # Mirrors roadmap formula: MIN(remaining, CEILING(ABS(remaining/weeks_left), 1))
        # Works for past ETA too: negative weeks_left → huge rate → MIN caps at remaining
        required_per_week = None
        eta_date = lookup_feature_eta_date(feature_eta_dates or {}, feature)
        if eta_date and remaining_count > 0:
            weeks_left = (eta_date - report_date).days / 7  # signed, can be negative
            if weeks_left == 0:
                required_per_week = remaining_count
            else:
                required_per_week = min(remaining_count, math.ceil(abs(remaining_count / weeks_left)))
        fdata.cell(row=fr_row, column=1, value=feature)
        fdata.cell(row=fr_row, column=2, value=required_per_week)
        fr_row += 1

        # ── write epic rows ───────────────────────────────────────────────────
        max_feature_epics = max(max_feature_epics, len(epic_stats))
        for seq, epic_key in enumerate(sorted(epic_stats.keys()), 1):
            s = epic_stats[epic_key]
            total = s["blank"] + s["in_progress"] + s["on_hold"] + s["done"] + s["rejected"]
            edata.cell(row=ed_row, column=1, value=feature)
            edata.cell(row=ed_row, column=2, value=epic_key if epic_key != "(no epic)" else "")
            edata.cell(row=ed_row, column=3, value=s["name"])
            edata.cell(row=ed_row, column=4, value=s["blank"])
            edata.cell(row=ed_row, column=5, value=s["in_progress"])
            edata.cell(row=ed_row, column=6, value=s["on_hold"])
            edata.cell(row=ed_row, column=7, value=s["done"])
            edata.cell(row=ed_row, column=8, value=s["rejected"])
            edata.cell(row=ed_row, column=9, value=total)
            edata.cell(row=ed_row, column=10, value=f"{feature}|{seq}")
            ed_row += 1

        # ── write weekly rows ─────────────────────────────────────────────────
        # Only show weeks where something was actually closed (done/rejected).
        # A week with only creations and no closures (e.g. a task created back
        # in 2017 under a long-running feature, still open) doesn't get its own
        # row — it would otherwise surface years with no real activity to report.
        all_week_keys = set(by_week.keys())
        if not all_week_keys:
            continue
        max_feature_weeks = max(max_feature_weeks, len(all_week_keys))
        empty_counts = {"tasks_done": 0, "tasks_rejected": 0, "bugs_done": 0, "bugs_rejected": 0}

        def _week_start(wk):
            yr, wn = wk
            try:
                return dt.date.fromisocalendar(yr, wn, 1)
            except Exception:
                return report_date

        # Sort by actual calendar date (not the bare week number) so cross-year
        # data — e.g. week 52 of last year vs. week 52 of this year — orders
        # correctly, with the current week always on top.
        for seq, wk in enumerate(sorted(all_week_keys, key=_week_start, reverse=True), 1):
            _, wn = wk
            counts = by_week.get(wk, empty_counts)
            total = counts["tasks_done"] + counts["tasks_rejected"] + counts["bugs_done"] + counts["bugs_rejected"]
            created = created_by_week.get(wk, 0)
            week_start_date = _week_start(wk)
            wdata.cell(row=wd_row, column=1, value=feature)
            wdata.cell(row=wd_row, column=2, value=wn)
            wdata.cell(row=wd_row, column=3, value=week_start_date)
            wdata.cell(row=wd_row, column=4, value=counts["tasks_done"])
            wdata.cell(row=wd_row, column=5, value=counts["tasks_rejected"])
            wdata.cell(row=wd_row, column=6, value=counts["bugs_done"])
            wdata.cell(row=wd_row, column=7, value=counts["bugs_rejected"])
            wdata.cell(row=wd_row, column=8, value=total)
            wdata.cell(row=wd_row, column=9, value=created)
            wdata.cell(row=wd_row, column=10, value=f"{feature}|{seq}")
            wd_row += 1

    last_ed_row = ed_row - 1
    last_fr_row = fr_row - 1
    last_wd_row = wd_row - 1

    # ── Layout constants ──────────────────────────────────────────────────────
    # Left table  (weekly):  cols A–J  (1–10)
    # Delimiter col:         K         (11)   width 4px
    # Right table (epics):  cols L–T  (12–20)
    #
    # Row 1: "Feature name" label (A1)
    # Row 2: feature dropdown (A2), "Required to close per week for ETA:" (F2), value (G2)
    # Row 3: blank spacer
    # Row 4: group sub-headers (weekly) + "Status by epic" italic label (epic side)
    # Row 5: column headers for both tables
    # Row 6+: data

    REQ_ROW        = 2
    REQ_LABEL_COL  = 6   # F
    REQ_VAL_COL    = 7   # G
    GROUP_ROW      = 4
    HEADER_ROW     = 5
    DATA_START     = 6
    DELIM_COL      = 11  # K
    EPIC_COL       = 12  # L — first epic table column

    no_wrap_left = Alignment(vertical="top", horizontal="left", wrap_text=False)

    # ── Required row (F4 / G4) ────────────────────────────────────────────────
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFFF8CC")
    cell = sheet.cell(row=REQ_ROW, column=REQ_LABEL_COL, value="Required")
    cell.font = header_font
    cell.alignment = no_wrap_left
    if last_fr_row >= 2:
        req_formula = f'=IFERROR(INDEX(\'_feature_req\'!$B$2:$B${last_fr_row},MATCH($A$2,\'_feature_req\'!$A$2:$A${last_fr_row},0)),"")'
        cell = sheet.cell(row=REQ_ROW, column=REQ_VAL_COL, value=req_formula)
        cell.font = base_font
        cell.alignment = left_top_alignment
        cell.fill = yellow_fill

    # ── Weekly group sub-header row (row 6) ───────────────────────────────────
    # Col A-B: "Weekly progress" italic spanning the two date/week cols
    cell = sheet.cell(row=GROUP_ROW, column=1, value="Weekly progress")
    cell.font = italic_font
    cell.alignment = no_wrap_left

    # Groups: Tasks (C-D), Bugs (E-F), Task+Bugs (G-J)
    weekly_groups = [(3, 4, "Tasks"), (5, 6, "Bugs"), (7, 10, "Task+Bugs")]
    for start_col, end_col, label in weekly_groups:
        cell = sheet.cell(row=GROUP_ROW, column=start_col, value=label)
        cell.font = italic_font
        cell.alignment = top_wrap_alignment
        sheet.merge_cells(start_row=GROUP_ROW, start_column=start_col, end_row=GROUP_ROW, end_column=end_col)
        sheet.cell(row=GROUP_ROW, column=start_col).border = Border(right=medium_side)

    # ── Epic group label (row 6, col L) ──────────────────────────────────────
    cell = sheet.cell(row=GROUP_ROW, column=EPIC_COL, value="Status by epic")
    cell.font = italic_font
    cell.alignment = no_wrap_left

    # ── Weekly column headers (row 7) — no fill, blue borders at section ends ─
    weekly_col_headers = ["Week start", "Week", "Done", "Rejected", "Done", "Rejected", "Closed", "Created", "Net", "Scope"]
    CLOSED_COL = 7  # G — highlighted with yellow fill throughout
    weekly_border_right = {2, 4, 6, 10}  # B, D, F, J — section boundaries
    for ci, h in enumerate(weekly_col_headers, 1):
        cell = sheet.cell(row=HEADER_ROW, column=ci, value=h)
        cell.font = header_font
        cell.alignment = top_wrap_alignment
        if ci == CLOSED_COL:
            cell.fill = yellow_fill
        if ci in weekly_border_right:
            cell.border = Border(right=medium_side)

    # ── Epic column headers (row 7, starting col L) ──────────────────────────
    epic_col_headers = ["Epic name", "Open", "Progress", "On Hold", "Done", "Rejected", "Total", "Progress %", "Link"]
    epic_border_right = {1, 7}  # after Epic name (L), after Total (R) — 1-based within epic block
    for ci, h in enumerate(epic_col_headers, 1):
        cell = sheet.cell(row=HEADER_ROW, column=EPIC_COL + ci - 1, value=h)
        cell.font = header_font
        cell.alignment = top_wrap_alignment
        if ci in epic_border_right:
            cell.border = Border(right=medium_side)

    # ── Epic data rows (MATCH/INDEX from _epic_data) ──────────────────────────
    def ed_range(letter):
        return f"'_epic_data'!${letter}$2:${letter}${last_ed_row}"

    data_rows = max(max_feature_epics, max_feature_weeks)

    for r in range(DATA_START, DATA_START + max_feature_epics):
        k = r - DATA_START + 1
        match_ed = f'MATCH($A$2&"|"&{k},{ed_range("J")},0)'
        iferr_ed = lambda letter, m=match_ed: f'=IFERROR(INDEX({ed_range(letter)},{m}),"")'

        # L: Epic name (col C in _epic_data), M-R: counts, S: Progress, T: Link
        # _epic_data: C=EpicName, D=Blank, E=InProgress, F=OnHold, G=Done, H=Rejected, I=Total
        col_map = {
            EPIC_COL:     ("C", False),   # Epic name
            EPIC_COL + 1: ("D", False),   # Open (blank)
            EPIC_COL + 2: ("E", False),   # In Progress
            EPIC_COL + 3: ("F", False),   # On Hold
            EPIC_COL + 4: ("G", False),   # Done
            EPIC_COL + 5: ("H", False),   # Rejected
            EPIC_COL + 6: ("I", False),   # Total
        }
        for col_idx, (ed_letter, _) in col_map.items():
            cell = sheet.cell(row=r, column=col_idx, value=iferr_ed(ed_letter))
            cell.font = base_font
            cell.alignment = left_top_alignment if col_idx == EPIC_COL else top_wrap_alignment
            offset = col_idx - EPIC_COL + 1
            if offset in epic_border_right:
                cell.border = Border(right=medium_side)

        # Progress bar: same style as summary "Scope done" column
        from openpyxl.utils import get_column_letter as gcl
        done_ref = gcl(EPIC_COL + 4) + str(r)      # P col (Done)
        rej_ref  = gcl(EPIC_COL + 5) + str(r)      # Q col (Rejected)
        tot_ref  = gcl(EPIC_COL + 6) + str(r)      # R col (Total)
        progress_formula = (
            f'=IFERROR(IF({tot_ref}=0,"",'
            f'TEXT(MIN(1,MAX(0,({done_ref}+{rej_ref})/{tot_ref})),"0%")&CHAR(10)&'
            f'REPT("█",ROUND(MIN(1,MAX(0,({done_ref}+{rej_ref})/{tot_ref}))*22,0))&'
            f'REPT("░",22-ROUND(MIN(1,MAX(0,({done_ref}+{rej_ref})/{tot_ref}))*22,0))),"")'
        )
        cell = sheet.cell(row=r, column=EPIC_COL + 7, value=progress_formula)
        cell.font = Font(name="Arial", size=10, color="FF5EC87A")
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Link: HYPERLINK to Jira
        link_formula = (
            f'=IFERROR(IF(INDEX({ed_range("B")},{match_ed})="",'
            f'"",HYPERLINK("https://{DEFAULT_HOST}/browse/"'
            f'&INDEX({ed_range("B")},{match_ed}),INDEX({ed_range("B")},{match_ed}))),"")'
        )
        cell = sheet.cell(row=r, column=EPIC_COL + 8, value=link_formula)
        cell.font = link_font
        cell.alignment = top_wrap_alignment

    # ── Weekly data rows (MATCH/INDEX from _weekly_data) ─────────────────────
    def wd_range(letter):
        return f"'_weekly_data'!${letter}$2:${letter}${last_wd_row}"

    for vis_row in range(DATA_START, DATA_START + max_feature_weeks):
        k = vis_row - DATA_START + 1
        match_wd = f'MATCH($A$2&"|"&{k},{wd_range("J")},0)'
        iferr_wd = lambda letter, m=match_wd: f'=IFERROR(INDEX({wd_range(letter)},{m}),"")'

        weekly_formulas = {
            1: iferr_wd("C"),   # WeekStart date
            2: iferr_wd("B"),   # Week number
            3: iferr_wd("D"),   # Tasks Done
            4: iferr_wd("E"),   # Tasks Rejected
            5: iferr_wd("F"),   # Bugs Done
            6: iferr_wd("G"),   # Bugs Rejected
            7: iferr_wd("H"),   # Closed (total done+rejected)
            8: iferr_wd("I"),   # Created
            9: f'=IF(A{vis_row}="","",H{vis_row}-G{vis_row})',    # Net = Created - Closed
            10: f'=IF(A{vis_row}="","",IF(I{vis_row}>=0,"🚀 Growing","🔥 Burning"))',
        }
        for col_idx, formula in weekly_formulas.items():
            cell = sheet.cell(row=vis_row, column=col_idx, value=formula)
            cell.font = base_font
            cell.alignment = left_top_alignment if col_idx <= 2 else top_wrap_alignment
            if col_idx == 1:
                cell.number_format = "dd.mmm.yy"
            if col_idx == 9:
                cell.number_format = "\\+0;\\-0;0"
            if col_idx == CLOSED_COL:
                cell.fill = yellow_fill
            if col_idx in weekly_border_right:
                cell.border = Border(right=medium_side)

    last_vis_row = DATA_START + max(data_rows, 1) - 1

    # Scope conditional formatting (col J)
    try:
        from openpyxl.formatting.rule import Rule
        from openpyxl.styles.differential import DifferentialStyle
        scope_range = f"J{DATA_START}:J{last_vis_row}"
        sheet.conditional_formatting.add(scope_range, Rule(
            type="formula",
            formula=[f'ISNUMBER(SEARCH("Burning",J{DATA_START}))'],
            dxf=DifferentialStyle(fill=light_green_fill),
        ))
        sheet.conditional_formatting.add(scope_range, Rule(
            type="formula",
            formula=[f'ISNUMBER(SEARCH("Growing",J{DATA_START}))'],
            dxf=DifferentialStyle(fill=light_red_fill),
        ))
    except Exception:
        pass

    sheet.auto_filter.ref = f"A{HEADER_ROW}:J{last_vis_row}"
    sheet.freeze_panes = f"A{DATA_START}"
    auto_fit_row_heights(sheet, 1, last_vis_row, min_height=14, line_height=14)

    # Delimiter col K: 4 Excel units wide
    sheet.column_dimensions["K"].width = 4

    from openpyxl.utils import get_column_letter
    col_widths = {
        1:  pixels_to_excel_width(100),   # A: Week start
        2:  10.17,   # B: Week
        3:  10.17,   # C: Tasks Done
        4:  10.17,   # D: Tasks Rejected
        5:  10.17,   # E: Bugs Done
        6:  10.17,   # F: Bugs Rejected
        7:  10.17,   # G: Closed
        8:  10.17,   # H: Created
        9:  10.17,   # I: Net
        10: pixels_to_excel_width(95),    # J: Scope
        # K delimiter set above
        12: pixels_to_excel_width(200),   # L: Epic name
        13: 10.17,   # M: Open
        14: 10.17,   # N: In Progress
        15: 10.17,   # O: On Hold
        16: 10.17,   # P: Done
        17: 10.17,   # Q: Rejected
        18: 10.17,   # R: Total
        19: 31,   # S: Progress %
        20: pixels_to_excel_width(90),    # T: Link
    }
    for col_num, width in col_widths.items():
        sheet.column_dimensions[get_column_letter(col_num)].width = width

    return sheet


def build_summary_sheet(sheet, rows, expected_tasks_per_week, feature_eta_dates, report_date, feature_styles):
    base_font = Font(name="Arial", size=10)
    header_font = Font(name="Arial", size=10, bold=True)
    link_font = Font(name="Arial", size=10, color="FF0563C1")
    top_wrap_alignment = Alignment(vertical="top", wrap_text=True)
    middle_alignment = Alignment(vertical="center")
    left_middle_alignment = Alignment(vertical="center", horizontal="left")
    top_left_alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
    no_wrap_left_alignment = Alignment(vertical="top", horizontal="left", wrap_text=False)
    percent_headers = {
        "Δ TTM, %",
        "Δ, %",
        "Δ ETA > 1d, %",
    }
    summary_headers = [
        "Feature",
        "Expected pace",
        "ETA",
        "Actual pace",
        "ETA from actual",
        "Required",
        "Scope done",
        "Delivered",
        "Δ ETA, mo",
        "Estimated",
        "Actual",
        "Hold, mo",
        "Δ, %",
        "Created",
        "Done",
        "Rejected",
        "Remaining",
        "Δ ETA > 1d",
        "Δ ETA > 1d, %",
        "Epics",
    ]

    sheet["A1"] = "Updated at"
    sheet["A1"].font = header_font
    sheet["A1"].alignment = middle_alignment
    _now = dt.datetime.now()
    sheet["A2"] = _now.strftime("%-d %b %H:%M") if hasattr(_now, "strftime") else report_date
    sheet["A2"].font = base_font
    sheet["A2"].alignment = top_left_alignment

    # Row 1: "Updated at" label. Row 2: the date, alone. Row 3: group sub-headers
    # (A3 is blank — "Feature" has no group label). Row 4: column headers. Row 5+: data.
    GROUP_ROW = 3
    HEADER_ROW = 4
    DATA_START = 5

    group_font = Font(name="Arial", size=10, italic=True)
    group_labels = [
        (2, "Pace - tasks per week (done+rejected)"),
        (7, "Status"),
        (10, "TTM, mo"),
        (14, "Totals tasks"),
        (18, "Tasks Δ ETA"),
        (20, "Epics"),
    ]
    for col_idx, label in group_labels:
        cell = sheet.cell(row=GROUP_ROW, column=col_idx, value=label)
        cell.font = group_font
        cell.alignment = no_wrap_left_alignment
    for col_idx in range(2, len(summary_headers) + 1):
        cell = sheet.cell(row=GROUP_ROW, column=col_idx)
        cell.alignment = Alignment(vertical="bottom", horizontal="left", wrap_text=False)

    for idx, header in enumerate(summary_headers, 1):
        cell = sheet.cell(row=HEADER_ROW, column=idx, value=header)
        cell.font = header_font
        cell.alignment = top_wrap_alignment

    medium_side = Side(style="thin", color="FF0072C6")
    section_boundary_cols = {1, 6, 9, 13, 17, 19}
    for idx in range(1, len(summary_headers) + 1):
        cell = sheet.cell(row=HEADER_ROW, column=idx)
        left = cell.border.left
        right = cell.border.right
        if idx in section_boundary_cols:
            right = medium_side
        if (idx - 1) in section_boundary_cols:
            left = medium_side
        cell.border = Border(left=left, right=right, top=cell.border.top, bottom=cell.border.bottom)

    grouped = {}
    for row in rows:
        feature = str(row.get("Feature") or "").strip()
        if not feature:
            continue
        grouped.setdefault(feature, {})
        issue_key = row.get("Link")
        if not issue_key:
            continue
        issue_group = grouped[feature].setdefault(issue_key, [])
        issue_group.append(row)

    widths = {header: len(header) for header in summary_headers}
    summary_rows = []
    for feature in sorted(grouped.keys(), key=str.lower):
        issue_rows_map = grouped[feature]
        created_count = 0
        done_count = 0
        rejected_count = 0
        remaining_count = 0
        in_progress_count = 0
        overdue_done_count = 0
        earliest_start = None
        earliest_done_start = None
        earliest_done_end = None
        earliest_created = None
        feature_active_work_dates = set()
        feature_active_weeks = set()
        done_last_4_weeks_by_week = {}
        created_last_4_weeks_by_week = {}
        epic_keys = set()
        latest_done_date = None

        for issue_key, issue_rows_list in issue_rows_map.items():
            epic_keys |= {str(item.get("Epic") or "").strip() for item in issue_rows_list if str(item.get("Epic") or "").strip()}
            current_row = current_issue_row(issue_rows_list)
            current_status = current_row.get("Status") or ""
            created_count += 1
            created_date = current_row.get("Created date")
            if created_date:
                if earliest_created is None or created_date < earliest_created:
                    earliest_created = created_date
                if in_last_n_weeks(created_date, report_date, weeks=4):
                    week_key = created_date.isocalendar()[:2]
                    created_last_4_weeks_by_week[week_key] = created_last_4_weeks_by_week.get(week_key, 0) + 1

            starts = [item.get("Start") for item in issue_rows_list if item.get("Start")]
            if starts:
                issue_start = min(starts)
                if earliest_start is None or issue_start < earliest_start:
                    earliest_start = issue_start

            if current_status == "rejected":
                rejected_count += 1
            elif current_status == "done":
                done_count += 1
                done_rows = [item for item in issue_rows_list if item.get("Status") == "done"]
                done_row = max(done_rows, key=row_effective_date) if done_rows else current_row
                done_start = done_row.get("Start")
                done_date = done_row.get("End")
                if done_start and done_date:
                    feature_active_work_dates |= business_day_dates(done_start, done_date)
                    feature_active_weeks |= {
                        day.isocalendar()[:2] for day in business_day_dates(done_start, done_date)
                    }
                if done_start and (earliest_done_start is None or done_start < earliest_done_start):
                    earliest_done_start = done_start
                if done_date and (earliest_done_end is None or done_date < earliest_done_end):
                    earliest_done_end = done_date
                if done_date and (latest_done_date is None or done_date > latest_done_date):
                    latest_done_date = done_date
                if done_date:
                    week_key = done_date.isocalendar()[:2]
                    done_last_4_weeks_by_week[week_key] = done_last_4_weeks_by_week.get(week_key, 0) + 1
                delta = done_row.get("Delta ETA")
                if isinstance(delta, (int, float)) and delta > 1:
                    overdue_done_count += 1
            elif current_status in {"", "in progress", "on hold"}:
                remaining_count += 1
                if current_status == "in progress":
                    in_progress_count += 1
            for row_item in issue_rows_list:
                row_status = row_item.get("Status")
                row_start = row_item.get("Start")
                row_end = row_item.get("End")
                if not (row_start and row_end):
                    continue
                segment_dates = business_day_dates(row_start, row_end)
                # feature_active_work_dates (Hold/TTM): every lifecycle row with both
                # dates represents real active work, whatever its status label ends up
                # being — issue_rows() sets a row's Start/End to the work burst leading
                # up to that transition (e.g. an "on hold" row's range is the work done
                # *before* it paused, not the pause itself; a "rejected" row's range is
                # the work spent before it was rejected). Only the blank/backlog row
                # has no dates at all, so don't filter this one by status.
                feature_active_work_dates |= segment_dates
                # feature_active_weeks (pace / "ETA from actual"): per spec, only weeks
                # with a done or in-progress row count as "active" here — including
                # on-hold/rejected weeks would understate the completion pace and push
                # the projected ETA out further than it should be.
                if row_status in {"done", "in progress"}:
                    feature_active_weeks |= {day.isocalendar()[:2] for day in segment_dates}

        feature_start = earliest_start or earliest_created
        eta_anchor_date = earliest_done_end or earliest_done_start
        avg_done_per_week_raw = (done_count / len(feature_active_weeks)) if feature_active_weeks else 0
        latest_done_weeks = sorted(done_last_4_weeks_by_week, reverse=True)[:4]
        expected_rate = lookup_expected_rate(expected_tasks_per_week, feature)
        eta_date = lookup_feature_eta_date(feature_eta_dates, feature)
        estimated_delivery_date = ""
        estimated_total_weeks = ""
        if eta_date:
            estimated_delivery_date = eta_date
            if not expected_rate and eta_anchor_date and eta_date > eta_anchor_date and created_count > 0:
                estimated_total_weeks = weeks_between(eta_anchor_date, eta_date)
                expected_rate = created_count / estimated_total_weeks if estimated_total_weeks else ""
        elif eta_anchor_date and expected_rate not in ("", None) and expected_rate > 0:
            estimated_total_weeks = created_count / expected_rate if expected_rate else ""
            estimated_delivery_date = eta_anchor_date + dt.timedelta(days=int(round(estimated_total_weeks * 7)))

        feature_done = (done_count + rejected_count) == created_count and created_count > 0
        # ETA is only ever a value someone actually set (a configured ETA date, or a
        # pace to project one from) — never invented from the delivery date itself.
        # A done feature with neither stays blank here; Delivered is the real answer.

        ttm_estimated = ""
        ttm_actual = ""
        def rounded_metric(value):
            if value in ("", None):
                return ""
            if isinstance(value, (int, float)):
                return int(round(value))
            return value

        avg_done_per_week = rounded_metric(avg_done_per_week_raw)
        eta_based_on_actual_pace = ""

        # Estimated TTM: calendar time from the first done task to the ETA — a plain
        # elapsed-time projection.
        # Actual TTM: distinct business days where ANY task had a real work burst —
        # unioned across every lifecycle row of every task in the feature — this is
        # `feature_active_work_dates`, built above. A row's Start/End is always the
        # work that happened *before* that row's terminal status, even "on hold" (the
        # pause itself is the gap between one row's End and the next row's Start, not
        # the row's own range) — see the comment at `feature_active_work_dates` above.
        # True idle gaps drop out on their own since no row contributes a date for
        # them, and parallel work on multiple tasks the same day only counts once, so
        # this measures real time-in-work rather than calendar span or summed task-days.
        BUSINESS_DAYS_PER_MONTH = 30.4375 * 5 / 7
        ttm_anchor_date = earliest_done_start or eta_anchor_date or earliest_done_end or latest_done_date
        if ttm_anchor_date and estimated_delivery_date and estimated_delivery_date > ttm_anchor_date:
            ttm_estimated = round((estimated_delivery_date - ttm_anchor_date).days / 30.4375, 1)
        delivery_date = latest_done_date if feature_done else ""
        if feature_done and feature_active_work_dates:
            ttm_actual = round(len(feature_active_work_dates) / BUSINESS_DAYS_PER_MONTH, 1)

        # Hold time: business days inside the feature's own working span (its first
        # to its last active day) where nothing was in progress or done — i.e. the
        # gap between calendar time and actual work time. This is what lets a
        # feature look "faster than estimated" (low Actual TTM) while still
        # delivering later than the ETA (positive Δ ETA) — the difference is time
        # spent paused, not time spent working.
        hold_months = ""
        if feature_done and feature_active_work_dates:
            full_span_dates = business_day_dates(min(feature_active_work_dates), max(feature_active_work_dates))
            hold_days = len(full_span_dates) - len(feature_active_work_dates)
            hold_months = round(hold_days / BUSINESS_DAYS_PER_MONTH, 1)

        # Calendar slip regardless of how much active work it took — delivered vs.
        # ETA, in months. Positive = missed the ETA, negative = delivered early.
        delta_eta_months = ""
        if delivery_date and estimated_delivery_date:
            delta_eta_months = round((delivery_date - estimated_delivery_date).days / 30.4375, 1)
        feature_status = next(
            (str(r.get("Feature status") or "").strip() for issue_rows_list in issue_rows_map.values() for r in issue_rows_list),
            "",
        )
        if eta_anchor_date and avg_done_per_week not in ("", None) and avg_done_per_week > 0 and created_count > 0 and feature_status != "on hold":
            actual_total_weeks = created_count / avg_done_per_week
            eta_based_on_actual_pace = eta_anchor_date + dt.timedelta(days=int(round(actual_total_weeks * 7)))

        active_scope_count = created_count if created_count else done_count + rejected_count + remaining_count
        completed_pct = ((done_count + rejected_count) / active_scope_count) if active_scope_count else 0

        required_done_per_week = ""
        if estimated_delivery_date and remaining_count > 0:
            weeks_left = (estimated_delivery_date - report_date).days / 7
            if weeks_left <= 0:
                required_done_per_week = remaining_count
            else:
                required_done_per_week = min(remaining_count, math.ceil(remaining_count / weeks_left))

        overdue_done_pct = (overdue_done_count / done_count) if done_count else 0
        epics_value = ", ".join(sorted(epic_keys))
        ttm_delta_pct = ""
        if ttm_estimated not in ("", None) and ttm_actual not in ("", None) and ttm_estimated > 0:
            ttm_delta_pct = (ttm_actual - ttm_estimated) / ttm_estimated

        def rounded(value):
            if value in ("", None):
                return ""
            if isinstance(value, (int, float)):
                return int(round(value))
            return value

        values = [
            feature,
            rounded(expected_rate),
            estimated_delivery_date,
            avg_done_per_week,
            eta_based_on_actual_pace,
            rounded(required_done_per_week),
            completed_pct,
            delivery_date,
            delta_eta_months,
            ttm_estimated,
            ttm_actual,
            hold_months,
            ttm_delta_pct,
            created_count,
            done_count,
            rejected_count,
            remaining_count,
            overdue_done_count,
            overdue_done_pct,
            epics_value,
        ]
        summary_rows.append(
            {
                "feature": feature,
                "values": values,
                "epic_keys": epic_keys,
                "ttm_delta_pct": ttm_delta_pct,
                "delta_eta_months": delta_eta_months,
                "remaining_count": remaining_count,
                "done_count": done_count,
                "created_count": created_count,
                "in_progress_count": in_progress_count,
                "epics_value": epics_value,
                "projected_delivery": estimated_delivery_date,
                "delivery_date": delivery_date,
                "eta_date": estimated_delivery_date,
                "eta_based_on_actual_pace": eta_based_on_actual_pace,
                "progress_value": completed_pct,
            }
        )

    def summary_sort_key(item):
        delivery_date = item["delivery_date"]
        eta_date = item["eta_date"]
        eta_actual = item["eta_based_on_actual_pace"]
        feature_name = item["feature"].lower()
        if delivery_date not in ("", None):
            return (0, delivery_date, feature_name)
        projected = [d for d in [eta_date, eta_actual] if d not in ("", None)]
        if projected:
            return (1, min(projected), feature_name)
        return (2, dt.date.max, feature_name)

    # Attach the sort key to each item so the sheet re-sort can reuse it
    # without having to re-derive it from cell values (which can lose type info).
    for item in summary_rows:
        item["_sort_key"] = summary_sort_key(item)

    dated_summary_rows = [item for item in summary_rows if item["_sort_key"][0] < 2]
    blank_summary_rows  = [item for item in summary_rows if item["_sort_key"][0] == 2]

    dated_summary_rows.sort(key=lambda item: item["_sort_key"])
    blank_summary_rows.sort(key=lambda item: item["feature"].lower())
    summary_rows = dated_summary_rows + blank_summary_rows

    row_sort_keys = {}  # sheet row number → pre-computed sort key
    row_idx = DATA_START
    for item in summary_rows:
        feature = item["feature"]
        values = item["values"]
        epic_keys = item["epic_keys"]
        epics_value = item["epics_value"]
        ttm_delta_pct = item["ttm_delta_pct"]
        delta_eta_months = item["delta_eta_months"]
        remaining_count = item["remaining_count"]
        done_count = item["done_count"]
        created_count = item["created_count"]
        in_progress_count = item["in_progress_count"]
        progress_value = item["progress_value"]

        for col_idx, (header, value) in enumerate(zip(summary_headers, values), 1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value if value != "" else None)
            cell.font = base_font
            cell.alignment = top_wrap_alignment
            if header == "Epics" and epics_value:
                cell.hyperlink = epic_list_link(epic_keys)
                cell.font = link_font
            if header in {"ETA", "ETA from actual", "Delivered"} and value:
                cell.number_format = "dd.mmm.yy"
            elif header in {"Δ ETA, mo", "Hold, mo", "TTM estimated, mo", "TTM actual, mo"} and value not in ("", None):
                cell.number_format = "0.0"
            elif header in percent_headers and value not in ("", None):
                set_numeric_format(cell, value, percent=True)
            elif isinstance(value, (int, float)):
                set_numeric_format(cell, value)
            if header == "Scope done":
                created_ref = sheet.cell(row=row_idx, column=summary_headers.index("Created") + 1).coordinate
                done_ref = sheet.cell(row=row_idx, column=summary_headers.index("Done") + 1).coordinate
                rejected_ref = sheet.cell(row=row_idx, column=summary_headers.index("Rejected") + 1).coordinate
                cell.value = (
                    f'=TEXT(MIN(1,MAX(0,(({done_ref}+{rejected_ref})/{created_ref}))),"0%")&CHAR(10)&'
                    f'REPT("█",ROUND(MIN(1,MAX(0,(({done_ref}+{rejected_ref})/{created_ref})))*22,0))&'
                    f'REPT("░",22-ROUND(MIN(1,MAX(0,(({done_ref}+{rejected_ref})/{created_ref})))*22,0))'
                )
                cell.font = Font(name="Arial", size=10, color="FF5EC87A")
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            elif header == "Δ, %" and ttm_delta_pct not in ("", None):
                font = copy(cell.font)
                font.color = "FFB71C1C" if ttm_delta_pct > 0 else "FF2E7D32"
                cell.font = font
            elif header == "Δ ETA, mo" and delta_eta_months not in ("", None):
                font = copy(cell.font)
                font.color = "FFB71C1C" if delta_eta_months > 0 else "FF2E7D32"
                cell.font = font

            left = cell.border.left
            right = cell.border.right
            if col_idx in section_boundary_cols:
                right = medium_side
            if (col_idx - 1) in section_boundary_cols:
                left = medium_side
            cell.border = Border(left=left, right=right, top=cell.border.top, bottom=cell.border.bottom)

            if True:
                if remaining_count == 0 and (done_count > 0 or created_count > 0):
                    cell.fill = PatternFill(fill_type="solid", fgColor="FFF0F0F0")
                elif in_progress_count > 0:
                    cell.fill = PatternFill(fill_type="solid", fgColor="FFFFF8CC")
                    font = copy(cell.font)
                    font.italic = True
                    cell.font = font

            display_value = value
            if isinstance(display_value, dt.date):
                display_value = display_value.strftime("%d.%b.%y")
            elif header in percent_headers and isinstance(display_value, (int, float)):
                display_value = f"{display_value:.0%}"
            widths[header] = max(widths[header], len("" if display_value in (None, "") else str(display_value)))

        row_sort_keys[row_idx] = item["_sort_key"]
        row_idx += 1

    sheet.freeze_panes = f"B{DATA_START}"  # freeze rows above the data AND column A (Feature)
    sheet.auto_filter.ref = f"A{HEADER_ROW}:{sheet.cell(row=HEADER_ROW, column=len(summary_headers)).column_letter}{max(row_idx - 1, HEADER_ROW)}"
    fixed_summary_widths = {
        "Feature": 34,
        "Epics": pixels_to_excel_width(179),
        "Scope done": 31,
    }
    for col_idx, header in enumerate(summary_headers, 1):
        width = fixed_summary_widths.get(header, 13.67)
        sheet.column_dimensions[sheet.cell(row=HEADER_ROW, column=col_idx).column_letter].width = width
    # Reorder the written summary rows from the actual sheet values so the
    # final workbook always matches the intended visible-date sort.
    summary_row_end = row_idx - 1
    if summary_row_end >= DATA_START:
        row_snaps = []
        for source_row in range(DATA_START, summary_row_end + 1):
            cells = []
            for col_idx in range(1, len(summary_headers) + 1):
                cell = sheet.cell(source_row, col_idx)
                cells.append(
                    {
                        "value": cell.value,
                        "style": copy(cell._style) if cell.has_style else None,
                        "number_format": cell.number_format,
                        "font": copy(cell.font),
                        "fill": copy(cell.fill),
                        "border": copy(cell.border),
                        "alignment": copy(cell.alignment),
                        "protection": copy(cell.protection),
                        "hyperlink": copy(cell.hyperlink) if cell.hyperlink else None,
                        "comment": copy(cell.comment) if cell.comment else None,
                    }
                )
            row_snaps.append(
                {
                    "source_row": source_row,
                    "sort_key": row_sort_keys.get(source_row, (2, dt.date.max, "")),
                    "height": sheet.row_dimensions[source_row].height,
                    "cells": cells,
                }
            )

        sorted_snaps = sorted(row_snaps, key=lambda snap: snap["sort_key"])
        for target_row, snap in enumerate(sorted_snaps, start=DATA_START):
            for col_idx, cell_data in enumerate(snap["cells"], start=1):
                dst = sheet.cell(target_row, col_idx)
                dst.value = cell_data["value"]
                if cell_data["style"] is not None:
                    dst._style = copy(cell_data["style"])
                dst.number_format = cell_data["number_format"]
                dst.font = copy(cell_data["font"])
                dst.fill = copy(cell_data["fill"])
                dst.border = copy(cell_data["border"])
                dst.alignment = copy(cell_data["alignment"])
                dst.protection = copy(cell_data["protection"])
                dst._hyperlink = copy(cell_data["hyperlink"]) if cell_data["hyperlink"] else None
                dst.comment = copy(cell_data["comment"]) if cell_data["comment"] else None
        auto_fit_row_heights(sheet, DATA_START, summary_row_end, min_height=14, line_height=14)
        # Leave rows 1-4 without an explicit height so both Excel and
        # Google Sheets auto-size them to fit the wrapped header text.


def read_existing_rows(path):
    workbook_path = Path(path)
    if not workbook_path.exists():
        return []
    workbook = load_workbook(workbook_path)
    sheet = workbook["tasks"] if "tasks" in workbook.sheetnames else workbook.active
    headers = [sheet.cell(1, col_idx).value for col_idx in range(1, sheet.max_column + 1)]
    rows = []
    for row_idx in range(2, sheet.max_row + 1):
        item = {headers[col_idx - 1]: sheet.cell(row_idx, col_idx).value for col_idx in range(1, sheet.max_column + 1)}
        rows.append(item)
    return rows


def normalize_existing_rows(existing_rows):
    rows = []
    for row in existing_rows:
        link = row.get("Link", "") or ""
        key = link.split("/")[-1] if "/" in link else link

        def parse_number(value):
            if value in ("", None):
                return ""
            try:
                return float(value)
            except Exception:
                return value

        rows.append(
            {
                "Feature": row.get("Feature", ""),
                "Epic": row.get("Epic", ""),
                "Substream": row.get("Substream", ""),
                "Task type": row.get("Task type", ""),
                "Task": row.get("Task", ""),
                "Status": row.get("Status", ""),
                "Start": row.get("Start").date() if isinstance(row.get("Start"), dt.datetime) else (row.get("Start") if isinstance(row.get("Start"), dt.date) else parse_date(row.get("Start", ""))),
                "End": row.get("End").date() if isinstance(row.get("End"), dt.datetime) else (row.get("End") if isinstance(row.get("End"), dt.date) else parse_date(row.get("End", ""))),
                "Done week": row.get("Done week", row.get("Week Num", "")),
                "ETA": parse_number(row.get("ETA", "")),
                "Days in Work": parse_number(row.get("Days in Work", "")),
                "Delta ETA": parse_number(row.get("Delta ETA", "")),
                "Link": key,
                "Created date": row.get("Created date").date() if isinstance(row.get("Created date"), dt.datetime) else (row.get("Created date") if isinstance(row.get("Created date"), dt.date) else parse_date(row.get("Created date", ""))),
                "Created week": row.get("Created week", ""),
                "Feature status": row.get("Feature status", ""),
            }
        )
    return rows


def merge_rows(existing_rows, fresh_rows):
    def identity(row):
        return (
            row["Link"],
            row["Task"],
            row["Status"],
            row["Start"],
            row["End"],
            row["Substream"],
        )

    preserved_map = {
        identity(row): row
        for row in existing_rows
        if row["Status"] in {"done", "rejected"}
    }
    merged = []
    seen = set()
    for row in fresh_rows:
        row_id = identity(row)
        if row_id in seen:
            continue
        seen.add(row_id)
        if row["Status"] in {"done", "rejected"} and row_id in preserved_map:
            merged.append(preserved_map[row_id])
        else:
            merged.append(row)
    return merged


def chunked(values, size):
    values = list(values or [])
    for index in range(0, len(values), max(1, size)):
        yield values[index : index + max(1, size)]


def scope_signature(include_values, exclude_values):
    return {
        "include": list(include_values or []),
        "exclude": list(exclude_values or []),
    }


def sanitize_expected_tasks_per_week(values):
    cleaned = {}
    for key, value in (values or {}).items():
        feature = str(key).strip()
        if not feature:
            continue
        try:
            number = float(value)
        except Exception:
            continue
        if number > 0:
            cleaned[feature] = number
    return cleaned


def main():
    parser = argparse.ArgumentParser(description="Generate a Jira Excel report.")
    parser.add_argument("--include", action="append", default=[], help="Required include keyword for epic summaries. Repeatable.")
    parser.add_argument("--include-file", help="Path to saved roadmap settings JSON.")
    parser.add_argument("--exclude", action="append", default=[], help="Exclude keyword for epic summaries. Repeatable.")
    parser.add_argument("--output", default=OUT_DEFAULT, help="Output xlsx path.")
    parser.add_argument("--state", default=STATE_DEFAULT, help="State file path.")
    parser.add_argument("--snapshot", default=str(SNAPSHOT_DEFAULT), help="Raw Jira snapshot JSON path.")
    parser.add_argument("--from-cache", action="store_true", help="Build from the raw Jira snapshot JSON instead of calling Jira.")
    parser.add_argument("--fresh", action="store_true", help="Build a fresh report without merging existing workbook rows.")
    parser.add_argument("--debug", action="store_true", help="Print verbose debug diagnostics.")
    parser.add_argument("--feature-filter", default="", help="Only process features matching this text.")
    parser.add_argument("--feature-filter-all", action="store_true", help="Use substring match for --feature-filter (default is exact match).")
    parser.add_argument("--update", action="store_true", help="Update mode: refresh only open/in-progress/on-hold tasks from the existing report.")
    parser.add_argument("--new-features", default="", help="Comma-separated list of new keyword(s) to full-fetch and add during --update.")
    args = parser.parse_args()
    global DEBUG, PROJECT_KEYS, DONE_STATUSES, DONE_STATUSES_LOWER
    DEBUG = args.debug

    # Prefer the local settings file; if it doesn't exist yet, seed it from the template.
    state_path = Path(args.state)
    if state_path == _SETTINGS_TEMPLATE and not _SETTINGS_LOCAL.exists() and _SETTINGS_TEMPLATE.exists():
        _SETTINGS_LOCAL.write_text(_SETTINGS_TEMPLATE.read_text())
        state_path = _SETTINGS_LOCAL
    elif state_path == _SETTINGS_TEMPLATE and _SETTINGS_LOCAL.exists():
        state_path = _SETTINGS_LOCAL
    args.state = str(state_path)

    state = {}
    if os.path.exists(args.state):
        with open(args.state, "r", encoding="utf-8") as f:
            state = json.load(f)
    # Normalise new `features` list → legacy flat keys so the rest of the
    # code keeps working unchanged.
    original_features_list = state.get("features") or []
    if original_features_list:
        features_list = original_features_list
        state["include"] = [f["keyword"] for f in features_list if f.get("keyword")]
        if "expected_tasks_per_week" not in state:
            state["expected_tasks_per_week"] = {
                f["keyword"]: f["expected_pace"]
                for f in features_list
                if f.get("keyword") and f.get("expected_pace") is not None
            }
        if "feature_eta_dates" not in state:
            state["feature_eta_dates"] = {
                f["keyword"]: f["eta"]
                for f in features_list
                if f.get("keyword") and f.get("eta")
            }
    spec = load_report_spec()
    include_values = []
    if args.include_file and os.path.exists(args.include_file):
        with open(args.include_file, "r", encoding="utf-8") as f:
            include_values = json.load(f).get("include", [])
    if args.include:
        include_values = list(dict.fromkeys(args.include + include_values))
    if not include_values:
        include_values = state.get("include", []) or ([spec["feature_keyword"]] if spec.get("feature_keyword") else [])
    exclude = args.exclude or state.get("exclude", []) or spec.get("exclude_keywords", [])
    output = args.output or state.get("output") or str(current_output_path())
    if not include_values:
        raise SystemExit("Missing include keywords and no saved state found.")
    if args.feature_filter:
        needle = normalize_keyword(args.feature_filter)
        if args.feature_filter_all:
            include_values = [k for k in include_values if needle in normalize_keyword(k)]
        else:
            include_values = [k for k in include_values if normalize_keyword(k) == needle]
        if not include_values:
            raise SystemExit(
                f"No features matched {'(substring)' if args.feature_filter_all else '(exact)'}: {args.feature_filter!r}"
            )
        print(f"Feature filter applied — updating: {', '.join(include_values)}")
    say_debug(f"DEBUG settings include keywords: {', '.join(include_values)}")
    say_debug(f"DEBUG settings exclude keywords: {', '.join(exclude) if exclude else '(none)'}")
    output_path = Path(output)
    yearly_default_output = current_output_path()
    if not output_path.is_absolute():
        output_path = (Path.cwd() / output_path).resolve()
    yearly_default_output_abs = (Path.cwd() / yearly_default_output).resolve()
    if (
        not args.output
        or output_path.name.startswith("roadmap ")
        or output_path.name in {"roadmap.xlsx", "jira_report.xlsx"}
    ):
        output_path = yearly_default_output_abs
    drive_folder = state.get("drive_folder")
    google_client_secrets = resolve_project_path(state.get("google_client_secrets"))
    local_only = state.get("local_only", True)
    update_time = state.get("update_time", "08:00")
    update_timezone = state.get("update_timezone", "UTC")
    PROJECT_KEYS = [str(k).strip().upper() for k in (state.get("project_keys") or []) if k]
    DONE_STATUSES = [str(s).strip() for s in (state.get("done_statuses") or DONE_STATUSES) if str(s).strip()]
    DONE_STATUSES_LOWER = {s.lower() for s in DONE_STATUSES}
    expected_tasks_per_week = sanitize_expected_tasks_per_week(state.get("expected_tasks_per_week"))
    feature_eta_dates = sanitize_feature_eta_dates(state.get("feature_eta_dates"))
    current_scope = scope_signature(include_values, exclude)
    previous_scope = state.get("_auto_generated", {}).get("last_generated_scope") or state.get("last_generated_scope")
    if args.from_cache or args.update:
        epic_name_field_ids = cached_field_ids(state, "epic_name_field_ids", ["customfield_10011"])
        eta_field_ids = cached_field_ids(state, "eta_field_ids", DEFAULT_ETA_FIELD_IDS)
        _cached_link = state.get("_auto_generated", {}).get("epic_link_field_id") or state.get("epic_link_field_id")
        epic_link_field_id = str(_cached_link) if _cached_link else "customfield_10014"
    else:
        epic_name_field_ids = detect_epic_name_field_ids()
        eta_field_ids = detect_eta_field_ids()
        epic_link_field_id = detect_epic_link_field_id()
    # Rebuild `features` list from canonical keyword order + any eta/pace data,
    # so the saved file stays in the tidy one-keyword-per-entry format.
    # When a feature filter is active, include_values is a subset — merge
    # updated entries back into the full list so non-filtered features are preserved.
    eta_map = {str(k).strip(): v for k, v in (feature_eta_dates or {}).items()}
    pace_map = {str(k).strip(): v for k, v in (expected_tasks_per_week or {}).items()}
    updated_entries = {}
    for kw in include_values:
        entry = {"keyword": kw}
        if kw in eta_map:
            entry["eta"] = eta_map[kw]
        if kw in pace_map:
            entry["expected_pace"] = pace_map[kw]
        updated_entries[kw] = entry
    if args.feature_filter and original_features_list:
        # Filter was active — preserve full features list, only update matched entries
        features_out = []
        for orig in original_features_list:
            kw = (orig.get("keyword") or "").strip()
            features_out.append(updated_entries.get(kw, orig))
    else:
        features_out = list(updated_entries.values())
    state = {
        "features": features_out,
        "exclude": exclude,
        "project_keys": PROJECT_KEYS,
        "output": str(output_path.relative_to(Path.cwd())) if output_path.is_relative_to(Path.cwd()) else str(output_path),
        "drive_folder": drive_folder,
        "google_client_secrets": make_project_relative(google_client_secrets),
        "local_only": local_only,
        "update_time": update_time,
        "update_timezone": update_timezone,
        "_auto_generated": {
            "epic_name_field_ids": epic_name_field_ids,
            "eta_field_ids": eta_field_ids,
            "epic_link_field_id": epic_link_field_id,
            "last_generated_scope": previous_scope,
        },
    }
    with open(args.state, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2)

    rows = []
    resolved_expected_tasks_per_week = dict(expected_tasks_per_week)
    cache_payload = {
        "generated_at": TODAY.isoformat(),
        "include": include_values,
        "exclude": exclude,
        "expected_tasks_per_week": expected_tasks_per_week,
        "feature_eta_dates": feature_eta_dates,
        "epics": [],
    }

    say_debug(f"DEBUG detected epic name fields: {', '.join(epic_name_field_ids)}")
    say_debug(f"DEBUG detected ETA fields: {', '.join(eta_field_ids)}")
    say_debug(f"DEBUG detected epic link field: {epic_link_field_id or '(none)'}")

    if args.update:
        # Task-level update: read existing report, find non-done tasks, fetch them fresh.
        # Nothing outside the matched tasks is touched.
        if not output_path.exists():
            raise SystemExit("No existing report found — run 'new' first.")

        keys_to_refresh = []
        update_existing = []
        rows_by_key = {}

        def _scan_existing(_set_msg):
            _set_msg("Reading existing report...")
            _rows = normalize_existing_rows(read_existing_rows(str(output_path)))
            if not _rows:
                raise SystemExit("Existing report is empty — run 'new' first.")
            update_existing.extend(_rows)

            for _row in _rows:
                _k = str(_row.get("Link") or "").strip()
                if _k:
                    rows_by_key.setdefault(_k, []).append(_row)

            if args.feature_filter:
                _needle = normalize_keyword(args.feature_filter)
                if args.feature_filter_all:
                    def _feat_match(_row):
                        return _needle in normalize_keyword(str(_row.get("Feature") or ""))
                else:
                    def _feat_match(_row):
                        return normalize_keyword(str(_row.get("Feature") or "")) == _needle
            else:
                def _feat_match(_row):
                    return True

            _total_keys = len(rows_by_key)
            for _i, (_k, _k_rows) in enumerate(rows_by_key.items()):
                if _i % 50 == 0:
                    _pct = int(_i / max(_total_keys, 1) * 100)
                    _set_msg(f"Scanning tasks {_pct}%...")
                if not any(_feat_match(_r) for _r in _k_rows):
                    continue
                _cur = current_issue_row(_k_rows)
                _status = str((_cur or {}).get("Status") or "").strip()
                if _status in ("", "in progress", "on hold"):
                    keys_to_refresh.append(_k)
            _set_msg("Scanning tasks 100%...")

        run_progress_spinner("Reading existing report...", _scan_existing)

        _has_new_features = bool([k.strip() for k in (args.new_features or "").split(",") if k.strip()])
        if not keys_to_refresh and not _has_new_features:
            say_done("All quiet — nothing changed. Check back later.")
            raise SystemExit(88)
        if keys_to_refresh:
            say_done(f"Found {len(keys_to_refresh)} task(s) to check for status changes.")
            if args.from_cache:
                # Dev mode: look up tasks from the snapshot instead of hitting Jira
                _cached = read_snapshot(args.snapshot)
                _cache_by_key = {
                    str(_ci.get("key") or "").strip(): _ci
                    for _entry in (_cached.get("epics") or [])
                    for _ci in (_entry.get("child_issues") or [])
                }
                _fetched_issues = [_cache_by_key[_k] for _k in keys_to_refresh if _k in _cache_by_key]
                say_done(f"Loaded {len(_fetched_issues)} task(s) from cache (dev mode).")
            else:
                _task_fields = ",".join(
                    ["summary", "status", "issuetype", "resolutiondate", "created", "project"] + eta_field_ids
                )
                _fetched_issues = []
                def _do_fetch(_set_msg):
                    _total = len(keys_to_refresh)
                    for _bi, _batch in enumerate(chunked(keys_to_refresh, 100), 1):
                        _pct = min(99, int(len(_fetched_issues) / max(_total, 1) * 100))
                        _set_msg(f"Fetching status updates {_pct}%...")
                        _jql = "issue in (" + ",".join(_batch) + ")"
                        _batch_issues = fetch_all_search(
                            _jql, _task_fields, expand="changelog",
                            context=f"update batch {_bi}",
                        )
                        _fetched_issues.extend(_batch_issues)
                    _set_msg(f"Fetching status updates 100%...")
                run_progress_spinner("Fetching status updates 0%...", _do_fetch)
            fresh_by_key = {str(_i.get("key") or "").strip(): _i for _i in _fetched_issues}

            # Re-generate rows for each refreshed task, track status changes
            fresh_rows_by_key = {}
            _changed = []
            _total_fresh = len(fresh_by_key)
            for _fi, (_key, _issue) in enumerate(fresh_by_key.items(), 1):
                _first = rows_by_key[_key][0]
                _feature = str(_first.get("Feature") or "")
                _epic_key = str(_first.get("Epic") or "")
                _substream = str(_first.get("Substream") or "")
                _new_rows = issue_rows(_issue, _feature, "", "", eta_field_ids, _epic_key)
                for _r in _new_rows:
                    if not _r.get("Substream"):
                        _r["Substream"] = _substream
                _old_status = str((current_issue_row(rows_by_key[_key]) or {}).get("Status") or "")
                _new_cur = current_issue_row(_new_rows)
                _new_status = str((_new_cur or {}).get("Status") or "") if _new_cur else ""
                if _old_status != _new_status:
                    _changed.append(_key)
                fresh_rows_by_key[_key] = _new_rows

            _new_feature_kwds_check = [k.strip() for k in (args.new_features or "").split(",") if k.strip()]
            if not _changed and not _new_feature_kwds_check:
                say_done("0 tasks updated their status")
                say_done("All quiet on the Jira front. Come back when someone actually does something.")
                raise SystemExit(88)
            elif _changed:
                say_done(f"{len(_changed)} tasks updated their status")

            # Splice updated task rows into existing — everything else stays untouched
            _final_rows = []
            _replaced = set()
            for _row in update_existing:
                _key = str(_row.get("Link") or "").strip()
                if _key in fresh_rows_by_key:
                    if _key not in _replaced:
                        _replaced.add(_key)
                        if fresh_rows_by_key[_key]:
                            _final_rows.extend(fresh_rows_by_key[_key])
                        # empty result means task is gone/rejected — drop it
                else:
                    _final_rows.append(_row)
            update_existing = _final_rows

        # Fetch and append rows for any new keywords added via --new-features
        _new_feature_kwds = [k.strip() for k in (args.new_features or "").split(",") if k.strip()]
        if _new_feature_kwds:
            _epic_fields = ",".join(["summary", "status", "issuetype", "project"] + epic_name_field_ids)
            _task_fields = ["summary", "status", "issuetype", "created", "updated", "resolutiondate"] + eta_field_ids
            if epic_link_field_id:
                _task_fields.append(epic_link_field_id)
            _existing_keys = {str(_r.get("Link") or "").strip() for _r in update_existing}
            for _nkwd in _new_feature_kwds:
                say_done(f"Fetching new feature: {_nkwd}")
                _nkwd_q = jql_quote(_nkwd.strip())
                _epics = run_spinner(
                    f"Collecting epics for '{_nkwd}'...",
                    lambda _q=_nkwd_q, _nk=_nkwd: fetch_all_search(
                        f'(summary ~ "{_q}" OR "Epic Name" ~ "{_q}") AND issuetype = Epic',
                        _epic_fields,
                        context=f"collecting epics for '{_nk}'",
                    ),
                )
                _matched_epics = [_e for _e in _epics if epic_matches_keyword(_e, _nkwd, epic_name_field_ids)]
                say_done(f"Found {len(_matched_epics)} epics for '{_nkwd}'")
                _nf_epic_by_key = {}
                for _e in _matched_epics:
                    _esummary = _e["fields"].get("summary", "")
                    _ename = get_epic_name(_e["fields"], epic_name_field_ids)
                    _flabel = pick_feature_label(_esummary, _ename, [_nkwd], _e["key"])
                    _nf_epic_by_key[_e["key"]] = {"epic": _e, "feature_label": _flabel, "child_issues": []}
                if not _nf_epic_by_key:
                    continue
                _nf_epic_keys = list(_nf_epic_by_key.keys())
                for _nf_batch in chunked(_nf_epic_keys, EPIC_BATCH_SIZE):
                    _epic_clause = " OR ".join(f'"Epic Link" = {_ek}' for _ek in _nf_batch)
                    _child_issues = fetch_all_search(
                        f"({_epic_clause})",
                        ",".join(_task_fields),
                        expand="changelog",
                        context=f"collecting tasks for new feature '{_nkwd}'",
                    )
                    for _ci in _child_issues:
                        _lev = _ci["fields"].get(epic_link_field_id) if epic_link_field_id else None
                        _lek = (_lev.get("key") or _lev.get("id") or "" if isinstance(_lev, dict) else str(_lev or "")).strip()
                        if _lek in _nf_epic_by_key:
                            _nf_epic_by_key[_lek]["child_issues"].append(_ci)
                for _nf_entry in _nf_epic_by_key.values():
                    _nf_epic = _nf_entry["epic"]
                    _nf_label = _nf_entry["feature_label"]
                    _nf_epic_summary = _nf_epic["fields"].get("summary", "")
                    _nf_epic_name = get_epic_name(_nf_epic["fields"], epic_name_field_ids)
                    for _ci in _nf_entry["child_issues"]:
                        _ci_key = str(_ci.get("key") or "").strip()
                        if _ci_key in _existing_keys:
                            continue  # task already in report, skip duplicate
                        _nf_rows = issue_rows(_ci, _nf_label, _nf_epic_summary, _nf_epic_name, eta_field_ids, _nf_epic["key"])
                        _nf_rows = [_r for _r in _nf_rows if row_belongs_in_current_report_year(_r, TODAY.year)]
                        update_existing.extend(_nf_rows)
                        if _ci_key:
                            _existing_keys.add(_ci_key)

        # Sort + annotate + write
        def _write_report(_set_msg):
            _set_msg("Updating report 25% — sorting rows...")
            def _to_date(_v):
                if isinstance(_v, dt.datetime):
                    return _v.date()
                return _v or dt.date.max
            update_existing.sort(key=lambda _r: (
                0 if _r.get("Start") else 1 if _r.get("End") else 2,
                _to_date(_r.get("Start")),
                _to_date(_r.get("End")),
                (_r.get("Substream") or "").lower(),
            ))
            _set_msg("Updating report 50% — annotating statuses...")
            update_existing[:] = annotate_feature_status(update_existing)
            _set_msg("Updating report 75% — building file...")
            build_xlsx(
                update_existing,
                str(output_path),
                expected_tasks_per_week=resolved_expected_tasks_per_week,
                feature_eta_dates=feature_eta_dates,
                report_date=TODAY,
            )
            _set_msg("Updating report 100%...")
        run_progress_spinner("Updating report 0%...", _write_report)
        say_done(f"Wrote {output_path} — {len(update_existing)} rows, updated at {dt.datetime.now().strftime('%-d %b %H:%M')}")
        raise SystemExit(88)  # Tell launcher: update done, handle upload and schedule line

    def collect_tasks_from_jira(set_message):
        # Features marked "done" in the existing report don't need re-fetching.
        # When a feature filter is active (direct --feature-filter without --update),
        # non-done non-filtered features pass through so the report stays complete.
        if not args.fresh and output_path.exists():
            existing_for_skip = normalize_existing_rows(read_existing_rows(str(output_path)))
            annotated_for_skip = annotate_feature_status(existing_for_skip)
            done_features = {
                str(r.get("Feature") or "").strip()
                for r in annotated_for_skip
                if r.get("Feature status") == "done"
            }
            skipped_rows = [r for r in annotated_for_skip if str(r.get("Feature") or "").strip() in done_features]
            if args.feature_filter:
                # Pass through rows for features that are neither done nor being updated
                include_values_set = {normalize_keyword(v) for v in include_values}
                passthrough_rows = [
                    r for r in annotated_for_skip
                    if str(r.get("Feature") or "").strip() not in done_features
                    and normalize_keyword(str(r.get("Feature") or "").strip()) not in include_values_set
                ]
                skipped_rows.extend(passthrough_rows)
        else:
            done_features = set()
            skipped_rows = []

        epics = []
        seen_keys = set()
        epic_fields = ",".join(["summary", "status", "issuetype", "project"] + epic_name_field_ids)
        for include in include_values:
            include_text = (include or "").strip()
            if include_text in done_features:
                say_done(f"Skipping '{include_text}' — already done")
                continue
            include_quoted = jql_quote(include_text)
            exclude_clauses = []
            for value in exclude:
                quoted = jql_quote(value)
                exclude_clauses.append(f'summary !~ "{quoted}"')
                exclude_clauses.append(f'"Epic Name" !~ "{quoted}"')
            project_name = all_project_name(include_text)
            if is_epic_key_keyword(include_text):
                found = run_spinner(
                    f"Collecting epics for '{include_text}'...",
                    lambda: [jget(f"{BASE}/issue/{include_text.upper()}?fields={epic_fields}", context=f"collecting epic {include_text}")],
                )
            elif project_name:
                found = run_spinner(
                    f"Collecting epics for '{include_text}'...",
                    lambda: fetch_all_search(
                        " AND ".join(
                            [
                                f'project = "{project_name}"',
                                "issuetype = Epic",
                            ] + exclude_clauses
                        ),
                        epic_fields,
                        context=f"collecting epics for '{include_text}'",
                    ),
                )
            else:
                found = run_spinner(
                    f"Collecting epics for '{include_text}'...",
                    lambda: fetch_all_search(
                        " AND ".join(
                            [
                                f'(summary ~ "{include_quoted}" OR "Epic Name" ~ "{include_quoted}")',
                                "issuetype = Epic",
                            ] + exclude_clauses
                        ),
                        epic_fields,
                        context=f"collecting epics for '{include_text}'",
                    ),
                )
            matched = []
            for epic in found:
                if not epic_matches_keyword(epic, include_text, epic_name_field_ids):
                    continue
                matched.append(epic)
                if epic["key"] not in seen_keys:
                    seen_keys.add(epic["key"])
                    epics.append(epic)
            say_done(f"Found {len(matched)} epics for '{include_text}'")
            if len(matched) > MAX_EPICS_PER_KEYWORD:
                raise SystemExit(
                    f"Woah... {len(matched)} epics is too many for '{include_text}'. "
                    "Try a more specific keyword so we do not pull half of Jira by accident."
                )

        epics[:] = [
            epic
            for epic in epics
            if not epic_matches_any_keyword(epic, exclude, epic_name_field_ids)
        ]
        say_debug(f"DEBUG epics after exclude filter: {len(epics)}")
        epic_entries = []
        epic_by_key = {}
        for epic in epics:
            epic_summary = epic["fields"].get("summary", "")
            epic_name = get_epic_name(epic["fields"], epic_name_field_ids)
            feature_label, source_include = pick_feature_label(epic_summary, epic_name, include_values, epic["key"], return_source=True)
            source_rate = lookup_expected_rate(expected_tasks_per_week, source_include)
            if source_rate not in ("", None):
                resolved_expected_tasks_per_week[feature_label] = source_rate
            entry = {
                "epic": epic,
                "feature_label": feature_label,
                "source_include": source_include,
                "child_issues": [],
            }
            epic_entries.append(entry)
            epic_by_key[epic["key"]] = entry

        epic_keys = [entry["epic"]["key"] for entry in epic_entries]
        task_fields = ["summary", "status", "issuetype", "created", "updated", "resolutiondate"] + eta_field_ids
        if epic_link_field_id:
            task_fields.append(epic_link_field_id)

        total_batches = max(1, math.ceil(len(epic_keys) / EPIC_BATCH_SIZE))
        say_debug(f"DEBUG epic entries: {len(epic_entries)}")
        say_debug(f"DEBUG epic batches: {total_batches}")
        for batch_index, epic_batch in enumerate(chunked(epic_keys, EPIC_BATCH_SIZE), 1):
            pct = min(100, int((batch_index / total_batches) * 100))
            set_message(f"Collecting tasks for report batch {batch_index}/{total_batches} ({pct}%)...")
            epic_clause = " OR ".join(f'"Epic Link" = {epic_key}' for epic_key in epic_batch)
            skip_statuses = ", ".join(f'"{s}"' for s in DONE_STATUSES) + ', "Rejected"'
            task_jql = f'({epic_clause}) AND status NOT IN ({skip_statuses})' if not args.fresh else f'({epic_clause})'
            child_issues = fetch_all_search(
                task_jql,
                ",".join(task_fields),
                expand="changelog",
                context=f"collecting task batch {batch_index}/{total_batches} for epics {', '.join(epic_batch)}",
            )
            for issue in child_issues:
                linked_epic_key = ""
                if epic_link_field_id:
                    linked_epic_value = issue["fields"].get(epic_link_field_id)
                    if isinstance(linked_epic_value, dict):
                        linked_epic_key = linked_epic_value.get("key") or linked_epic_value.get("id") or ""
                    else:
                        linked_epic_key = str(linked_epic_value or "").strip()
                if linked_epic_key and linked_epic_key in epic_by_key:
                    epic_by_key[linked_epic_key]["child_issues"].append(issue)
            time.sleep(REQUEST_PAUSE_SECONDS)
        say_debug(
            "DEBUG child issues per epic: "
            + ", ".join(f"{entry['epic']['key']}={len(entry['child_issues'])}" for entry in epic_entries[:10])
            + (" ..." if len(epic_entries) > 10 else "")
        )

        for idx, entry in enumerate(epic_entries, 1):
            pct = min(100, int((idx / max(len(epic_entries), 1)) * 100))
            set_message(f"Collecting tasks for report {idx}/{len(epic_entries)} ({pct}%)...")
            epic = entry["epic"]
            epic_summary = epic["fields"].get("summary", "")
            epic_name = get_epic_name(epic["fields"], epic_name_field_ids)
            child_issues = entry["child_issues"]
            cache_payload["epics"].append(entry)
            project_counts = {}
            for issue in child_issues:
                project_key = ((issue.get("fields") or {}).get("project") or {}).get("key") or ""
                project_key = str(project_key).strip().upper() or "(none)"
                project_counts[project_key] = project_counts.get(project_key, 0) + 1
            say_debug(
                "DEBUG child project counts for "
                f"{epic['key']}: "
                + ", ".join(f"{key}={value}" for key, value in sorted(project_counts.items()))
            )
            epic_rows = []
            for issue in child_issues:
                issue_task_rows = issue_rows(issue, entry["feature_label"], epic_summary, epic_name, eta_field_ids, epic["key"])
                if not issue_task_rows:
                    continue
                epic_rows.extend(issue_task_rows)
            feature_rows = [row for row in epic_rows if row_belongs_in_current_report_year(row, TODAY.year)]
            feature_has_current_year_activity = rows_have_feature_report_activity(feature_rows, TODAY.year)
            if feature_has_current_year_activity:
                rows.extend(feature_rows)
            say_debug(
                f"DEBUG epic {entry['epic']['key']} feature='{entry['feature_label']}' "
                f"child_issues={len(child_issues)} rows_added={len(feature_rows)} current_year={feature_has_current_year_activity}"
            )
        rows.extend(skipped_rows)
        say_debug(f"DEBUG total rows after Jira processing: {len(rows)}")

    def collect_tasks_from_cache(set_message):
        cached = read_snapshot(args.snapshot)
        if not cached:
            raise SystemExit(f"No raw snapshot found at {snapshot_path(args.snapshot)}")
        cached_epics = cached.get("epics", [])
        cached_rows = cached.get("rows") or []
        if not cached_epics and cached_rows:
            normalized = normalize_existing_rows(cached_rows)
            grouped = {}
            order = []
            for row in normalized:
                key = row.get("Feature") or ""
                grouped.setdefault(key, []).append(row)
                if key not in order:
                    order.append(key)
            for key in order:
                feature_rows = grouped.get(key, [])
                feature_rows = [row for row in feature_rows if row_belongs_in_current_report_year(row, TODAY.year)]
                if rows_have_feature_report_activity(feature_rows, TODAY.year):
                    rows.extend(feature_rows)
            say_debug(f"DEBUG cache rows loaded directly: {len(rows)}")
            return
        total = len(cached_epics)
        for idx, entry in enumerate(cached_epics, 1):
            pct = min(100, int((idx / max(total, 1)) * 100))
            set_message(f"Collecting tasks for report {idx}/{max(total, 1)} ({pct}%)...")
            epic = entry["epic"]
            feature_label = entry["feature_label"]
            epic_summary = epic["fields"].get("summary", "")
            epic_name = get_epic_name(epic["fields"], epic_name_field_ids)
            source_include = entry.get("source_include") or feature_label
            source_rate = lookup_expected_rate(expected_tasks_per_week, source_include)
            if source_rate not in ("", None):
                resolved_expected_tasks_per_week[feature_label] = source_rate
            epic_rows = []
            project_counts = {}
            for issue in entry.get("child_issues", []):
                project_key = ((issue.get("fields") or {}).get("project") or {}).get("key") or ""
                project_key = str(project_key).strip().upper() or "(none)"
                project_counts[project_key] = project_counts.get(project_key, 0) + 1
            say_debug(
                "DEBUG cache child project counts for "
                f"{epic['key']}: "
                + ", ".join(f"{key}={value}" for key, value in sorted(project_counts.items()))
            )
            for issue in entry.get("child_issues", []):
                issue_task_rows = issue_rows(issue, feature_label, epic_summary, epic_name, eta_field_ids, epic["key"])
                if not issue_task_rows:
                    continue
                epic_rows.extend(issue_task_rows)
            feature_rows = [row for row in epic_rows if row_belongs_in_current_report_year(row, TODAY.year)]
            feature_has_current_year_activity = rows_have_feature_report_activity(feature_rows, TODAY.year)
            if feature_has_current_year_activity:
                rows.extend(feature_rows)
            say_debug(
                f"DEBUG cache epic {epic['key']} feature='{feature_label}' "
                f"child_issues={len(entry.get('child_issues', []))} rows_added={len(feature_rows)} current_year={feature_has_current_year_activity}"
            )
        say_debug(f"DEBUG total rows after cache processing: {len(rows)}")

    if args.from_cache:
        run_progress_spinner("Collecting tasks for report 1/1 (0%)...", collect_tasks_from_cache)
    else:
        run_progress_spinner("Collecting tasks for report 1/1 (0%)...", collect_tasks_from_jira)
    say_done(f"Collected {len(rows)} rows from Jira")
    say_debug(f"DEBUG rows before merge/sort: {len(rows)}")

    if args.fresh:
        existing = []
    elif previous_scope and previous_scope != current_scope:
        say_done("Keyword scope changed. Reusing the existing workbook so done and rejected rows stay preserved.")
        existing = normalize_existing_rows(read_existing_rows(str(output_path)))
    else:
        existing = normalize_existing_rows(read_existing_rows(str(output_path)))
    def sort_rows():
        nonlocal rows
        rows = merge_rows(existing, rows)
        def _to_date(v):
            if isinstance(v, dt.datetime):
                return v.date()
            return v or dt.date.max
        rows.sort(key=lambda r: (
            0 if r["Start"] else 1 if r["End"] else 2,
            _to_date(r["Start"]),
            _to_date(r["End"]),
            (r["Substream"] or "").lower(),
        ))
        rows = annotate_feature_status(rows)
        say_debug(f"DEBUG rows after merge/sort/status annotation: {len(rows)}")

    def write_file():
        build_xlsx(
            rows,
            str(output_path),
            expected_tasks_per_week=resolved_expected_tasks_per_week,
            feature_eta_dates=feature_eta_dates,
            report_date=TODAY,
        )

    run_spinner("Sorting rows for report...", sort_rows)
    run_spinner("Writing the report file...", write_file)
    say_done(f"Wrote {output_path} with {len(rows)} rows")

    if not args.from_cache:
        write_snapshot(
            args.snapshot,
            {
                **cache_payload,
                "rows": rows,
                "resolved_expected_tasks_per_week": resolved_expected_tasks_per_week,
            },
        )

    state.setdefault("_auto_generated", {})["last_generated_scope"] = current_scope
    with open(args.state, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2)

    # Drive upload is handled by the launcher after this process exits


if __name__ == "__main__":
    try:
        main()
    except JiraAuthError as exc:
        sys.stderr.write(f"{exc}\n")
        raise SystemExit(86)
    except JiraNetworkError as exc:
        sys.stderr.write(f"{exc}\n")
        raise SystemExit(87)
    except KeyboardInterrupt:
        raise SystemExit(0)
